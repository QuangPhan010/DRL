import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io

class ExcelExporter:
    """
    Excel Report Exporter for Training Points.
    Generates a stylized excel file using openpyxl.
    """
    def export(self, data, parameters):
        # Styles definition
        title_font = Font(name="Calibri", size=16, bold=True, color="1F497D")
        subtitle_font = Font(name="Calibri", size=11, italic=True)
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Calibri", size=11)
        
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        
        thin_side = Side(border_style="thin", color="D3D3D3")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        is_audit = len(data) > 0 and 'username' in data[0] and 'action' in data[0]

        if is_audit:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Audit Log"
            ws.views.sheetView[0].showGridLines = True
            
            ws.append([])
            ws.cell(row=2, column=1, value="BÁO CÁO VẾT HỆ THỐNG (AUDIT LOG)").font = title_font
            ws.cell(row=3, column=1, value="Nhật ký chi tiết các thao tác, truy cập và thay đổi cấu hình").font = subtitle_font
            
            start_row = 5
            headers = ["STT", "Ngày giờ", "Tài khoản", "Vai trò", "Thao tác", "Đối tượng", "Giá trị trước", "Giá trị sau", "Địa chỉ IP"]
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = align_center
                cell.border = border
                
            for row_idx, item in enumerate(data, 1):
                r = start_row + row_idx
                ws.cell(row=r, column=1, value=row_idx).alignment = align_center
                ws.cell(row=r, column=2, value=item.get('created_at', '')).alignment = align_center
                ws.cell(row=r, column=3, value=item.get('username', '')).alignment = align_left
                ws.cell(row=r, column=4, value=item.get('role', '')).alignment = align_center
                ws.cell(row=r, column=5, value=item.get('action', '')).alignment = align_left
                ws.cell(row=r, column=6, value=item.get('entity_name', '')).alignment = align_left
                ws.cell(row=r, column=7, value=item.get('before_value', '')).alignment = align_left
                ws.cell(row=r, column=8, value=item.get('after_value', '')).alignment = align_left
                ws.cell(row=r, column=9, value=item.get('ip_address', '')).alignment = align_center
                
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=r, column=col_idx)
                    cell.font = data_font
                    cell.border = border
                    
            for col in ws.columns:
                max_len = 0
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                for cell in col[start_row-1:]:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
                
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

        is_activity = len(data) > 0 and 'activity_title' in data[0]

        if is_activity:
            wb = openpyxl.Workbook()
            default_sheet = wb.active
            is_first = True
            
            for act_idx, act_data in enumerate(data):
                title = act_data.get('activity_title', 'Hoạt động')
                for char in ['*', ':', '?', '/', '\\', '[', ']']:
                    title = title.replace(char, '')
                sheet_title = f"{act_idx+1}_{title[:25]}"
                
                if is_first:
                    ws = default_sheet
                    ws.title = sheet_title
                    is_first = False
                else:
                    ws = wb.create_sheet(title=sheet_title)
                    
                ws.views.sheetView[0].showGridLines = True
                
                ws.append([])
                ws.cell(row=2, column=1, value="BÁO CÁO CHI TIẾT HOẠT ĐỘNG NGOẠI KHÓA").font = title_font
                ws.cell(row=3, column=1, value=f"Hoạt động: {act_data.get('activity_title')}").font = Font(name="Calibri", size=12, bold=True, color="1F497D")
                ws.cell(row=4, column=1, value=f"Ngày tổ chức: {act_data.get('activity_date')}  |  Địa điểm: {act_data.get('activity_location')}").font = subtitle_font
                
                start_row = 6
                headers = ["STT", "MSSV", "Họ và tên", "Lớp", "Khoa", "Giờ Check-in", "Giờ Check-out", "Trạng thái"]
                
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=start_row, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = align_center
                    cell.border = border
                    
                for row_idx, p in enumerate(act_data.get('participants', []), 1):
                    r = start_row + row_idx
                    ws.cell(row=r, column=1, value=row_idx).alignment = align_center
                    ws.cell(row=r, column=2, value=p.get('student_id', '')).alignment = align_center
                    ws.cell(row=r, column=3, value=p.get('full_name', '')).alignment = align_left
                    ws.cell(row=r, column=4, value=p.get('class_name', '')).alignment = align_center
                    ws.cell(row=r, column=5, value=p.get('faculty', '')).alignment = align_left
                    ws.cell(row=r, column=6, value=p.get('checkin_time', '')).alignment = align_center
                    ws.cell(row=r, column=7, value=p.get('checkout_time', '')).alignment = align_center
                    
                    status_cell = ws.cell(row=r, column=8, value=p.get('status', ''))
                    status_cell.alignment = align_center
                    if p.get('status') == "Đầy đủ":
                        status_cell.font = Font(name="Calibri", size=11, bold=True, color="10B981")
                    else:
                        status_cell.font = Font(name="Calibri", size=11, bold=True, color="EF4444")
                        
                    for col_idx in range(1, len(headers) + 1):
                        cell = ws.cell(row=r, column=col_idx)
                        if col_idx != 8:
                            cell.font = data_font
                        cell.border = border
                        
                for col in ws.columns:
                    max_len = 0
                    col_letter = openpyxl.utils.get_column_letter(col[0].column)
                    for cell in col[start_row-1:]:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

        # Group data by class_name
        data_by_class = {}
        for item in data:
            c_name = item.get('class_name') or 'Khác'
            if c_name not in data_by_class:
                data_by_class[c_name] = []
            data_by_class[c_name].append(item)

        wb = openpyxl.Workbook()
        default_sheet = wb.active
        
        school_year = parameters.get('school_year', '')
        semester = parameters.get('semester', '')
        faculty = parameters.get('faculty', '')
        
        headers = [
            "STT", "MSSV", "Họ và tên", "Lớp", "Khoa", 
            "GPA", "Xếp loại học tập", "Điểm tự đánh giá", 
            "Điểm rèn luyện tổng", "Xếp loại rèn luyện", "Trạng thái"
        ]

        is_first = True
        # If no data at all, just keep one empty sheet
        if not data_by_class:
            default_sheet.title = "Trống"
            default_sheet.views.sheetView[0].showGridLines = True
            default_sheet.cell(row=2, column=1, value="Không có dữ liệu sinh viên").font = title_font
        else:
            for c_name, class_data in data_by_class.items():
                sheet_title = c_name[:30]
                if is_first:
                    ws = default_sheet
                    ws.title = sheet_title
                    is_first = False
                else:
                    ws = wb.create_sheet(title=sheet_title)
                
                # Show grid lines
                ws.views.sheetView[0].showGridLines = True
                
                # Title lines
                ws.append([]) # Row 1
                ws.cell(row=2, column=1, value="BÁO CÁO TỔNG HỢP ĐIỂM RÈN LUYỆN SINH VIÊN").font = title_font
                ws.cell(row=3, column=1, value=f"Năm học: {school_year} | Học kỳ: {semester}").font = subtitle_font
                
                filters = []
                if faculty:
                    filters.append(f"Khoa: {faculty}")
                filters.append(f"Lớp: {c_name}")
                ws.cell(row=4, column=1, value=" - ".join(filters)).font = subtitle_font
                start_row = 6
                
                # Headers Row
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=start_row, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = align_center
                    cell.border = border
                    
                # Write data rows
                for row_idx, item in enumerate(class_data, 1):
                    r = start_row + row_idx
                    ws.cell(row=r, column=1, value=row_idx).alignment = align_center
                    ws.cell(row=r, column=2, value=item.get('student_id', '')).alignment = align_center
                    ws.cell(row=r, column=3, value=item.get('full_name', '')).alignment = align_left
                    ws.cell(row=r, column=4, value=item.get('class_name', '')).alignment = align_center
                    ws.cell(row=r, column=5, value=item.get('faculty', '')).alignment = align_left
                    ws.cell(row=r, column=6, value=item.get('gpa', 0.0)).alignment = align_right
                    ws.cell(row=r, column=7, value=item.get('gpa_classification', '')).alignment = align_center
                    ws.cell(row=r, column=8, value=item.get('self_score', 0)).alignment = align_right
                    ws.cell(row=r, column=9, value=item.get('total_score', 0)).alignment = align_right
                    ws.cell(row=r, column=10, value=item.get('classification', '')).alignment = align_center
                    ws.cell(row=r, column=11, value=item.get('status', '')).alignment = align_center
                    
                    for col_idx in range(1, len(headers) + 1):
                        cell = ws.cell(row=r, column=col_idx)
                        cell.font = data_font
                        cell.border = border
                        
                # Auto-fit column widths
                for col in ws.columns:
                    max_len = 0
                    col_letter = openpyxl.utils.get_column_letter(col[0].column)
                    for cell in col[start_row-1:]:
                        if cell.value:
                            val_str = str(cell.value)
                            max_len = max(max_len, len(val_str))
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
        # Save to byte stream
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
