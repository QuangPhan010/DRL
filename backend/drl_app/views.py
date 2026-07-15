import math
import datetime
from django.utils import timezone
from django.db import transaction
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from django.shortcuts import get_object_or_404
from django.contrib.auth import authenticate
from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi

from .models import User, ClassInfo, Student, CriteriaSet, Criterion, GroupCriterion, SubItem, Evaluation, EvaluationDetail, Activity, ActivityParticipant, Organization, UserOrganization, ClassPosition, StudentClassPosition, ActivityCheckIn, ActivityCheckOut, ActivityAttendance, FraudDetection, AuditLog, ChangeRequest, ExternalActivity, EvidenceFile, EvidenceReview, FraudFlag, SystemConfig, Notification, Room
from .serializers import (
    UserSerializer, ClassInfoSerializer, StudentSerializer, CriteriaSetSerializer, CriterionSerializer,
    EvaluationSerializer, ActivitySerializer, ActivityParticipantSerializer,
    OrganizationSerializer, UserOrganizationSerializer, ClassPositionSerializer, StudentClassPositionSerializer,
    ActivityCheckInSerializer, ActivityCheckOutSerializer, ActivityAttendanceSerializer, FraudDetectionSerializer, AuditLogSerializer, ChangeRequestSerializer,
    ExternalActivitySerializer, EvidenceFileSerializer, EvidenceReviewSerializer, FraudFlagSerializer, SystemConfigSerializer,
    NotificationSerializer, RoomSerializer
)

def create_notification(user, title, message, type='system', level='info', action_url=None):
    try:
        notification = Notification.objects.create(
            user=user,
            title=title,
            message=message,
            type=type,
            level=level,
            action_url=action_url
        )
        try:
            from channels.layers import get_channel_layer
            from asgiref.sync import async_to_sync
            channel_layer = get_channel_layer()
            if channel_layer:
                async_to_sync(channel_layer.group_send)(
                    f"user_{user.id}",
                    {
                        "type": "send_notification",
                        "notification": {
                            "id": notification.id,
                            "user": user.id,
                            "title": notification.title,
                            "message": notification.message,
                            "unread": notification.unread,
                            "type": notification.type,
                            "level": notification.level,
                            "action_url": notification.action_url,
                            "created_at": notification.created_at.isoformat() if notification.created_at else None
                        }
                    }
                )
        except Exception as ws_err:
            print(f"Error broadcasting websocket notification: {ws_err}")
    except Exception as e:
        print(f"Error creating notification: {e}")

import base64
import numpy as np

FACE_MATCH_THRESHOLD = 0.60
FACE_PRESENTATION_THRESHOLD = 0.60

def extract_face_embedding_from_base64(base64_str):
    try:
        # Face recognition is optional because its native dependencies do not
        # support every Python version that can run the rest of the API.
        # Import lazily so a missing ML runtime cannot prevent Django starting.
        import cv2
        from deepface import DeepFace

        if ',' in base64_str:
            base64_str = base64_str.split(',')[1]
        img_data = base64.b64decode(base64_str)
        nparr = np.frombuffer(img_data, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        if img is None:
            return None, "Dữ liệu ảnh không hợp lệ."
        
        # Downscale image to max 480px width/height to speed up face detection and extraction significantly
        max_dim = 480
        h, w = img.shape[:2]
        if max(h, w) > max_dim:
            scale = max_dim / max(h, w)
            img = cv2.resize(img, (int(w * scale), int(h * scale)))
        
        # Heuristic for cartoon/anime image detection
        # Resize to 100x100 and count unique colors to detect flat shading of cartoon/illustrations
        resized = cv2.resize(img, (100, 100))
        unique_colors = len(np.unique(resized.reshape(-1, 3), axis=0))
        if unique_colors < 1000:
            return None, "Ảnh đại diện phải là ảnh thật có mặt của bạn. Không được sử dụng ảnh hoạt hình, ảnh vẽ hoặc ảnh đồ họa."

        # Extract embedding using DeepFace (using Facenet model)
        objs = DeepFace.represent(img_path=img, model_name="Facenet", enforce_detection=True)
        if len(objs) == 0:
            return None, "Không phát hiện được khuôn mặt nào."
        
        # If multiple faces, pick the largest one (which is closest to camera)
        if len(objs) > 1:
            objs = sorted(objs, key=lambda face: face["facial_area"]["w"] * face["facial_area"]["h"], reverse=True)
            
        return objs[0]["embedding"], None
    except ImportError as e:
        import traceback
        traceback.print_exc()
        return None, (
            "Tính năng nhận diện khuôn mặt chưa được cài đặt. "
            "Hãy cài các gói trong requirements.txt bằng Python 3.10-3.13. "
            f"Chi tiết lỗi: {str(e)}"
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        err_msg = str(e)
        if "Face could not be detected" in err_msg:
            return None, "Không phát hiện được khuôn mặt trong ảnh. Vui lòng chụp rõ mặt."
        return None, f"Lỗi xử lý khuôn mặt: {err_msg}"

def _face_similarity(reference, candidate):
    if not isinstance(reference, list) or not isinstance(candidate, list):
        return 0.0
    if len(reference) != len(candidate) or len(reference) == 0:
        return 0.0
    try:
        a = np.array([float(x) for x in reference])
        b = np.array([float(x) for x in candidate])
        if not np.all(np.isfinite(a)) or not np.all(np.isfinite(b)):
            return 0.0
        norm_a = np.linalg.norm(a)
        norm_b = np.linalg.norm(b)
        if norm_a == 0 or norm_b == 0:
            return 0.0
        similarity = np.dot(a, b) / (norm_a * norm_b)
        if not math.isfinite(similarity):
            return 0.0
        if similarity >= 1.0 - 1e-12:
            return 1.0
        return float(max(min(similarity, 1.0), 0.0))
    except Exception:
        return 0.0

def _verify_attendance_face(request):
    """Compare one live scan only with the authenticated user's avatar."""
    reference = request.user.avatar_embedding
    if not request.user.avatar or not reference:
        return None, Response(
            {'error': 'Bạn cần cập nhật ảnh đại diện có khuôn mặt trước khi điểm danh.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    face_image = request.data.get('faceImage') or request.data.get('face_image')
    if not face_image:
        return None, Response(
            {'error': 'Thiếu ảnh quét khuôn mặt.'},
            status=status.HTTP_400_BAD_REQUEST,
        )
        
    scan_embedding, error_msg = extract_face_embedding_from_base64(face_image)
    if error_msg:
        return None, Response(
            {'error': error_msg},
            status=status.HTTP_400_BAD_REQUEST,
        )

    similarity = _face_similarity(reference, scan_embedding)
    if similarity < FACE_MATCH_THRESHOLD:
        return None, Response(
            {
                'error': f'Khuôn mặt không khớp với ảnh đại diện (Độ tương đồng: {int(similarity * 100)}%).',
                'face_similarity': similarity,
            },
            status=status.HTTP_403_FORBIDDEN,
        )
    return {
        'similarity': similarity,
        'liveness': 1.0,
        'realness': 1.0,
    }, None


def _verify_attendance_location(request, activity, student):
    # GPS validation is completely disabled. Return mock zero values.
    return {
        'distance': 0.0,
        'accuracy': 0.0,
        'latitude': 0.0,
        'longitude': 0.0,
    }, None


def _can_review_activity_attendance(user, activity):
    if not user or not user.is_authenticated:
        return False

    roles = set(user.roles)
    if roles.intersection({'admin', 'student_affairs', 'academic_affairs'}):
        return True
    if activity.organizer and activity.organizer == user.full_name:
        return True
    if user.user_organizations.filter(organization__name=activity.organizer).exists():
        return True

    if activity.scope_type == 'club':
        return user.user_organizations.filter(
            organization__in=activity.allowed_clubs.all(),
            position__in=['Chủ nhiệm', 'Phó chủ nhiệm', 'Trưởng ban', 'Phụ trách'],
        ).exists()

    if activity.scope_type == 'class':
        allowed_classes = activity.allowed_classes.all()
        if user.role == 'advisor' and allowed_classes.filter(advisor=user).exists():
            return True
        student = getattr(user, 'student_profile', None)
        return bool(
            student
            and student.class_info_id
            and allowed_classes.filter(pk=student.class_info_id).exists()
            and roles.intersection({'class_monitor', 'advisor'})
        )

    return False


def _reserve_activity_participant(activity_id, student):
    """Create one participant without allowing concurrent overbooking."""
    with transaction.atomic():
        activity = Activity.objects.select_for_update().get(pk=activity_id)
        participant = ActivityParticipant.objects.filter(
            activity=activity,
            student=student,
        ).first()
        if participant:
            return participant, False, False
        if activity.participants.count() >= activity.max_participants:
            return None, False, True
        participant = ActivityParticipant.objects.create(
            activity=activity,
            student=student,
            status='registered',
        )
        return participant, True, False


def _activity_start_at(activity):
    start_time = activity.start_time or datetime.time.min
    starts_at = datetime.datetime.combine(activity.date, start_time)
    return timezone.make_aware(starts_at, timezone.get_current_timezone())


def _activity_schedule(activity):
    starts_at = _activity_start_at(activity)
    if activity.end_time:
        ends_at = timezone.make_aware(
            datetime.datetime.combine(activity.date, activity.end_time),
            timezone.get_current_timezone(),
        )
        if ends_at <= starts_at:
            ends_at += datetime.timedelta(days=1)
    else:
        duration = max(int(activity.duration_minutes or 180), 1)
        ends_at = starts_at + datetime.timedelta(minutes=duration)
    return starts_at, ends_at


class CriteriaAdminOrReadOnly(permissions.BasePermission):
    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return True
        return (
            request.user
            and request.user.is_authenticated
            and request.user.role in ('admin', 'student_affairs')
        )

# 1. Login View
@swagger_auto_schema(
    method='post',
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        required=['username', 'password'],
        properties={
            'username': openapi.Schema(type=openapi.TYPE_STRING),
            'password': openapi.Schema(type=openapi.TYPE_STRING)
        }
    ),
    responses={200: openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'token': openapi.Schema(type=openapi.TYPE_STRING),
            'user': openapi.Schema(type=openapi.TYPE_OBJECT)
        }
    )}
)
@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def login_view(request):
    username = request.data.get('username')
    password = request.data.get('password')
    
    from django.db.models import Q
    user_obj = User.objects.filter(Q(username=username) | Q(student_id=username)).first()
    if user_obj:
        user = authenticate(username=user_obj.username, password=password)
        if user:
            if not user.is_active:
                return Response({'error': 'Tài khoản đã bị đóng.'}, status=status.HTTP_403_FORBIDDEN)
            return Response({
                'token': f"mock-token-for-{user.username}",
                'user': UserSerializer(user).data,
                'is_first_login': user.is_first_login
            })
    return Response({'error': 'Invalid username or password'}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def change_password_view(request):
    username = request.data.get('username')
    new_password = request.data.get('password')
    if not username or not new_password:
        return Response({'error': 'Username and password are required'}, status=status.HTTP_400_BAD_REQUEST)
    
    from django.db.models import Q
    user = User.objects.filter(Q(username=username) | Q(student_id=username)).first()
    if not user:
        return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
        
    user.set_password(new_password)
    user.is_first_login = False
    user.plain_password = ""
    user.save()
    return Response({'message': 'Password changed successfully', 'user': UserSerializer(user).data})

@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def request_first_login_password(request):
    student_id = request.data.get('student_id')
    if not student_id:
        return Response({'error': 'Vui lòng cung cấp mã sinh viên (MSSV).'}, status=status.HTTP_400_BAD_REQUEST)
    
    student = Student.objects.filter(student_id__iexact=student_id).first()
    if not student:
        return Response({'error': 'Không tìm thấy sinh viên trong hệ thống.'}, status=status.HTTP_404_NOT_FOUND)
        
    user = student.user
    if not user:
        return Response({'error': 'Tài khoản sinh viên chưa được liên kết.'}, status=status.HTTP_400_BAD_REQUEST)
        
    if not user.is_first_login or user.last_login is not None:
        return Response({'error': 'Tài khoản đã đổi mật khẩu lần đầu.'}, status=status.HTTP_400_BAD_REQUEST)
        
    import random
    import string
    random_password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
    user.set_password(random_password)
    user.plain_password = random_password
    user.save()
    
    from django.core.mail import send_mail
    from django.conf import settings
    
    subject = "[ITC Point] Mật khẩu đăng nhập lần đầu"
    message = f"Chào {student.full_name},\n\nTài khoản ITC Point của bạn đã sẵn sàng.\nMật khẩu đăng nhập lần đầu của bạn là: {random_password}\n\nVui lòng đăng nhập và thực hiện đổi mật khẩu ngay sau đó."
    
    html_message = f"""
<div style="font-family: 'Segoe UI', Roboto, Helvetica, Arial, sans-serif; max-width: 600px; margin: 0 auto; border: 1px solid #e2e8f0; border-radius: 16px; overflow: hidden; box-shadow: 0 10px 25px -5px rgba(0,0,0,0.05), 0 8px 10px -6px rgba(0,0,0,0.05); background-color: #ffffff;">
  <!-- Header with gradient background -->
  <div style="background: linear-gradient(135deg, #3b82f6, #06b6d4); padding: 32px 24px; text-align: center; color: white;">
    <h1 style="margin: 0; font-size: 24px; font-weight: 800; letter-spacing: -0.5px;">ITC Point</h1>
    <p style="margin: 4px 0 0 0; font-size: 13px; color: rgba(255,255,255,0.85); font-weight: 500;">Hệ Thống Quản Lý Điểm Rèn Luyện Trực Tuyến</p>
  </div>
  
  <!-- Main content body -->
  <div style="padding: 40px 32px; background-color: #ffffff;">
    <h2 style="margin-top: 0; color: #1e293b; font-size: 20px; font-weight: 700; letter-spacing: -0.5px;">Chào {student.full_name},</h2>
    <p style="color: #475569; line-height: 1.6; font-size: 15px; margin-top: 12px;">Tài khoản đánh giá điểm rèn luyện trực tuyến của bạn đã được thiết lập thành công. Vui lòng sử dụng thông tin đăng nhập tạm thời dưới đây để đăng nhập lần đầu vào hệ thống:</p>
    
    <!-- Info Box -->
    <div style="background-color: #f8fafc; border-radius: 12px; padding: 20px; margin: 28px 0; border: 1px solid #f1f5f9; border-left: 4px solid #3b82f6;">
      <table style="width: 100%; border-collapse: collapse;">
        <tr>
          <td style="padding: 6px 0; font-size: 14px; color: #64748b; font-weight: 600; width: 130px; text-transform: uppercase; letter-spacing: 0.5px;">Tên đăng nhập:</td>
          <td style="padding: 6px 0; font-size: 15px; color: #0f172a; font-weight: 700; font-family: monospace;">{student.student_id}</td>
        </tr>
        <tr>
          <td style="padding: 6px 0; font-size: 14px; color: #64748b; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px;">Mật khẩu tạm:</td>
          <td style="padding: 6px 0; font-size: 16px; color: #ef4444; font-weight: 700; font-family: monospace; letter-spacing: 1px;">{random_password}</td>
        </tr>
      </table>
    </div>

    <!-- CTA Button -->
    <div style="text-align: center; margin: 36px 0 28px 0;">
      <a href="http://localhost:8080/login" style="background: linear-gradient(135deg, #3b82f6, #06b6d4); color: #ffffff; padding: 14px 32px; text-decoration: none; border-radius: 10px; font-weight: 700; font-size: 15px; display: inline-block; box-shadow: 0 10px 15px -3px rgba(59, 130, 246, 0.3); letter-spacing: -0.2px;">Đăng nhập hệ thống</a>
    </div>

    <!-- Important Notice -->
    <div style="background-color: #fef2f2; border-radius: 8px; padding: 14px 18px; border: 1px solid #fee2e2; margin-top: 24px;">
      <p style="color: #b91c1c; font-size: 13px; font-weight: 600; margin: 0; line-height: 1.5;">
        * Quan trọng: Đây là mật khẩu tự động tạm thời. Bạn bắt buộc phải đổi sang mật khẩu bảo mật của riêng mình ngay tại lần đầu đăng nhập.
      </p>
    </div>
  </div>
  
  <!-- Footer section -->
  <div style="background-color: #f8fafc; padding: 24px; text-align: center; border-top: 1px solid #f1f5f9; color: #94a3b8; font-size: 12px; line-height: 1.5;">
    <p style="margin: 0 0 6px 0; font-weight: 500;">Email này được hệ thống ITC Point gửi tự động.</p>
    <p style="margin: 0;">© 2026 ITC Point. All rights reserved.</p>
  </div>
</div>
"""
    
    try:
        send_mail(
            subject,
            message,
            getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@itcpoint.com'),
            [student.email],
            fail_silently=False,
            html_message=html_message,
        )
        return Response({'message': 'Mật khẩu đã được gửi về email của sinh viên thành công.'})
    except Exception as e:
        return Response({
            'message': 'Đã cập nhật mật khẩu hệ thống (Gửi mail thất bại, cấu hình SMTP chưa đúng).',
            'plain_password': random_password
        })

class RoomViewSet(viewsets.ModelViewSet):
    queryset = Room.objects.all()
    serializer_class = RoomSerializer
    permission_classes = [permissions.AllowAny]

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            class IsAdminUserRole(permissions.BasePermission):
                def has_permission(self, request, view):
                    return request.user.is_authenticated and request.user.role == 'admin'
            return [IsAdminUserRole()]
        return [permissions.AllowAny()]

    @action(detail=False, methods=['post'], url_path='check-availability')
    def check_availability(self, request):
        date_str = request.data.get('date')
        start_time_str = request.data.get('start_time')
        end_time_str = request.data.get('end_time')
        exclude_activity_id = request.data.get('exclude_activity_id')

        if not (date_str and start_time_str and end_time_str):
            return Response({'error': 'Vui lòng cung cấp đầy đủ thông tin: date, start_time, end_time.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            if isinstance(date_str, str):
                date_val = datetime.datetime.strptime(date_str.split('T')[0], '%Y-%m-%d').date()
            else:
                date_val = date_str

            # Normalize times (only take HH:MM, ignore seconds if present)
            st_clean = ':'.join(start_time_str.split(':')[:2])
            et_clean = ':'.join(end_time_str.split(':')[:2])

            start_time_val = datetime.datetime.strptime(st_clean, '%H:%M').time()
            end_time_val = datetime.datetime.strptime(et_clean, '%H:%M').time()
        except ValueError as e:
            return Response({'error': f'Định dạng ngày/giờ không hợp lệ: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        rooms = Room.objects.all()
        result = []

        for room in rooms:
            conflicts = Activity.objects.filter(
                room=room,
                date=date_val,
                start_time__lt=end_time_val,
                end_time__gt=start_time_val
            )
            
            if exclude_activity_id:
                try:
                    conflicts = conflicts.exclude(id=int(exclude_activity_id))
                except ValueError:
                    pass

            conflicting_activity = conflicts.first()

            if conflicting_activity:
                st_str = conflicting_activity.start_time.strftime('%H:%M') if conflicting_activity.start_time else ""
                et_str = conflicting_activity.end_time.strftime('%H:%M') if conflicting_activity.end_time else ""
                conflict_info = {
                    'title': conflicting_activity.title,
                    'time': f"{st_str} - {et_str}"
                }
                is_available = False
            else:
                conflict_info = None
                is_available = True

            result.append({
                'id': room.id,
                'name': room.name,
                'capacity': room.capacity,
                'location': room.location,
                'is_available': is_available,
                'conflict_info': conflict_info
            })

        return Response(result)

# 2. ClassInfo ViewSet
class ClassInfoViewSet(viewsets.ModelViewSet):
    serializer_class = ClassInfoSerializer
    permission_classes = [permissions.AllowAny] # In actual build, we can restrict by roles.

    def get_queryset(self):
        user = self.request.user
        if user.is_authenticated and user.role == 'advisor':
            return ClassInfo.objects.filter(advisor=user)
        
        advisor_id = self.request.query_params.get('advisorId') or self.request.query_params.get('advisor_id')
        if advisor_id:
            return ClassInfo.objects.filter(advisor_id=advisor_id)
            
        return ClassInfo.objects.all()

    @action(detail=True, methods=['post'], url_path='assign-advisor')
    def assign_advisor(self, request, pk=None):
        class_info = self.get_object_value(pk)
        advisor_id = request.data.get('advisorId') or request.data.get('advisor_id')
        if not advisor_id:
            class_info.advisor = None
            class_info.save()
            return Response(ClassInfoSerializer(class_info).data)
        
        advisor = get_object_or_404(User, id=advisor_id, role='advisor')
        class_info.advisor = advisor
        class_info.save()
        return Response(ClassInfoSerializer(class_info).data)

    @action(detail=True, methods=['post'], url_path='assign-monitor')
    def assign_monitor(self, request, pk=None):
        class_info = self.get_object_value(pk)
        student_id = request.data.get('studentId') or request.data.get('student_id')
        student = get_object_or_404(Student, student_id=student_id, class_info=class_info)
        
        # Demote previous monitor
        position, _ = ClassPosition.objects.get_or_create(name='Lớp trưởng')
        StudentClassPosition.objects.filter(class_info=class_info, position=position).delete()
        
        # Assign new monitor
        assigned_by = request.user if request.user.is_authenticated else None
        StudentClassPosition.objects.create(student=student, class_info=class_info, position=position, assigned_by=assigned_by)
        
        # Backwards compatibility: update roles
        User.objects.filter(student_id=student_id).update(role='class_monitor')
        
        return Response({'message': f'Student {student_id} is now the class monitor'})

    @action(detail=True, methods=['post'], url_path='assign-position')
    def assign_position(self, request, pk=None):
        class_info = self.get_object_value(pk)
        student_id = request.data.get('studentId') or request.data.get('student_id')
        position_name = request.data.get('positionName') or request.data.get('position_name')
        
        student = get_object_or_404(Student, student_id=student_id, class_info=class_info)
        position, _ = ClassPosition.objects.get_or_create(name=position_name)
        
        assigned_by = request.user if request.user.is_authenticated else None
        
        # Unique positions per class
        if position_name in ['Lớp trưởng', 'Lớp phó', 'Bí thư']:
            StudentClassPosition.objects.filter(class_info=class_info, position=position).delete()
            
            # Backwards compatibility with the static role
            if position_name == 'Lớp trưởng':
                class_students = Student.objects.filter(class_info=class_info)
                User.objects.filter(student_id__in=class_students.values_list('student_id', flat=True), role='class_monitor').update(role='student')
                User.objects.filter(student_id=student_id).update(role='class_monitor')

        StudentClassPosition.objects.update_or_create(
            student=student,
            class_info=class_info,
            position=position,
            defaults={'assigned_by': assigned_by}
        )
        
        return Response({'message': f'Student {student_id} has been assigned position {position_name}'})

    @action(detail=True, methods=['post'], url_path='revoke-position')
    def revoke_position(self, request, pk=None):
        class_info = self.get_object_value(pk)
        student_id = request.data.get('studentId') or request.data.get('student_id')
        position_name = request.data.get('positionName') or request.data.get('position_name')
        
        student = get_object_or_404(Student, student_id=student_id, class_info=class_info)
        
        StudentClassPosition.objects.filter(
            student=student,
            class_info=class_info,
            position__name=position_name
        ).delete()
        
        # If revoking Lớp trưởng, revert role to student for backwards compatibility
        if position_name == 'Lớp trưởng':
            User.objects.filter(student_id=student_id, role='class_monitor').update(role='student')
            
        return Response({'message': f'Revoked position {position_name} from student {student_id}'})

    def get_object_value(self, pk):
        return get_object_or_404(ClassInfo, pk=pk)

    @action(detail=True, methods=['get', 'post'], url_path='students')
    def students(self, request, pk=None):
        class_info = self.get_object_value(pk)
        if request.method == 'GET':
            students = Student.objects.filter(class_info=class_info)
            return Response(StudentSerializer(students, many=True).data)
        
        elif request.method == 'POST':
            student_id = request.data.get('studentId') or request.data.get('student_id')
            student = get_object_or_404(Student, student_id=student_id)
            student.class_info = class_info
            student.save()
            return Response(StudentSerializer(student).data)

    @action(detail=True, methods=['delete'], url_path='students/(?P<student_id>[^/.]+)')
    def remove_student(self, request, pk=None, student_id=None):
        class_info = self.get_object_value(pk)
        student = get_object_or_404(Student, student_id=student_id, class_info=class_info)
        student.class_info = None
        student.save()
        return Response({'message': 'Student removed from class successfully'})

# 3. Student ViewSet
class StudentViewSet(viewsets.ModelViewSet):
    serializer_class = StudentSerializer
    permission_classes = [permissions.AllowAny]

    def get_queryset(self):
        user = self.request.user
        queryset = Student.objects.all()
        if user.is_authenticated and user.role == 'advisor':
            queryset = queryset.filter(class_info__advisor=user)
            
        student_id = self.request.query_params.get('student_id') or self.request.query_params.get('studentId')
        if student_id:
            queryset = queryset.filter(student_id=student_id)
            
        class_name = self.request.query_params.get('class_name') or self.request.query_params.get('className')
        if class_name:
            queryset = queryset.filter(class_info__name=class_name)
            
        # Self-healing: auto-create missing User objects for Student objects
        orphans = queryset.filter(user__isnull=True)
        if orphans.exists():
            import string
            import random
            for s in orphans:
                username = s.student_id.lower()
                user_obj = User.objects.filter(username__iexact=username).first()
                if not user_obj:
                    user_obj = User.objects.filter(student_id=s.student_id).first()
                if not user_obj:
                    characters = string.ascii_letters + string.digits
                    random_password = ''.join(random.choice(characters) for i in range(8))
                    user_obj = User.objects.create_user(
                        username=username,
                        email=s.email,
                        password=random_password,
                        role='student',
                        full_name=s.full_name,
                        student_id=s.student_id,
                        is_first_login=True,
                        plain_password=random_password
                    )
                s.user = user_obj
                s.save()
            
        return queryset

    @action(detail=False, methods=['post'], url_path='import-excel')
    def import_excel(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'No file was submitted'}, status=status.HTTP_400_BAD_REQUEST)

        import openpyxl
        import random
        try:
            wb = openpyxl.load_workbook(file_obj, read_only=True)
            students_created = 0
            processed_sheets = 0
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                first_row = True
                header_map = {}
                
                # Check and parse sheet
                sheet_valid = True
                for row in sheet.iter_rows(values_only=True):
                    if first_row:
                        first_row = False
                        for idx, cell in enumerate(row):
                            if cell:
                                header_map[str(cell).strip().lower()] = idx
                        
                        # Validate headers
                        required_id_terms = ['mã sv', 'ma sv', 'student id', 'student_id', 'mã sinh viên', 'ma sinh vien', 'mssv', 'mã số sinh viên', 'ma so sinh vien']
                        required_name_terms = ['họ và tên', 'ho va ten', 'full name', 'fullname', 'tên', 'ten', 'họ tên', 'ho ten']
                        
                        has_id = any(term in header_map for term in required_id_terms)
                        has_name = any(term in header_map for term in required_name_terms)
                        
                        if not (has_id and has_name):
                            sheet_valid = False
                            break
                        processed_sheets += 1
                        continue
                    
                    if not sheet_valid:
                        break

                    # Skip completely empty rows
                    if not any(row):
                        continue

                    def get_val(names):
                        for name in names:
                            if name in header_map:
                                idx = header_map[name]
                                if idx < len(row):
                                    val = row[idx]
                                    return str(val).strip() if val is not None else ''
                        return ''

                    student_id = get_val(['mã sv', 'ma sv', 'student id', 'student_id', 'mã sinh viên', 'ma sinh vien', 'mssv', 'mã số sinh viên', 'ma so sinh vien'])
                    full_name = get_val(['họ và tên', 'ho va ten', 'full name', 'fullname', 'tên', 'ten', 'họ tên', 'ho ten'])
                    email = get_val(['email', 'thư điện tử', 'thu dien tu', 'hòm thư', 'hom thu'])
                    class_name = get_val(['lớp', 'lop', 'class', 'class name', 'class_name', 'lớp sinh hoạt', 'lop sinh hoat'])
                    faculty = get_val(['khoa', 'faculty', 'khoa đào tạo', 'khoa dao tao']) or 'Công nghệ Thông tin'
                    cohort = get_val(['khóa', 'khoa', 'cohort', 'niên khóa', 'nien khoa', 'niên khoá']) or 'K20'
                    gender = get_val(['giới tính', 'gioi tinh', 'gender', 'phái', 'phai']) or 'Nam'
                    phone = get_val(['số điện thoại', 'so dien thoai', 'phone', 'sđt', 'sdt', 'điện thoại', 'dien thoai'])

                    if not student_id or not full_name:
                        continue

                    # Auto-generate fallback email if missing
                    if not email:
                        email = f"{student_id.lower()}@stu.itc.edu.vn"

                    # Find or create class
                    from .models import ClassInfo
                    class_obj = None
                    if class_name:
                        class_obj, _ = ClassInfo.objects.get_or_create(
                            name=class_name,
                            defaults={'faculty': faculty, 'cohort': cohort}
                        )

                    student_obj = Student.objects.filter(student_id=student_id).first()
                    if not student_obj:
                        user_obj = User.objects.filter(student_id=student_id).first()
                        if not user_obj:
                            user_obj = User.objects.filter(username__iexact=student_id).first()
                        
                        if not user_obj:
                            username = student_id.lower()
                            random_password = generate_random_password()
                            user_obj = User.objects.create_user(
                                username=username,
                                email=email,
                                password=random_password,
                                role='student',
                                full_name=full_name,
                                student_id=student_id,
                                is_first_login=True,
                                plain_password=random_password
                            )
                        else:
                            # Update student_id on existing User if missing
                            if not user_obj.student_id:
                                user_obj.student_id = student_id
                                user_obj.save()

                        Student.objects.create(
                            user=user_obj,
                            student_id=student_id,
                            full_name=full_name,
                            email=email,
                            class_info=class_obj,
                            faculty=faculty,
                            cohort=cohort,
                            gender=gender,
                            phone=phone
                        )
                        students_created += 1
                    else:
                        student_obj.full_name = full_name
                        student_obj.email = email
                        if class_obj:
                            student_obj.class_info = class_obj
                        student_obj.faculty = faculty
                        student_obj.cohort = cohort
                        student_obj.gender = gender
                        if phone:
                            student_obj.phone = phone
                        student_obj.save()
                        
                        if student_obj.user:
                            student_obj.user.full_name = full_name
                            student_obj.user.email = email
                            student_obj.user.save()

            if processed_sheets == 0:
                return Response({
                    'error': "Không tìm thấy bất kỳ sheet nào có cấu trúc hợp lệ (chứa tiêu đề 'Mã SV' và 'Họ và tên'). Vui lòng kiểm tra lại file Excel."
                }, status=status.HTTP_400_BAD_REQUEST)

            return Response({'message': f'Nhập thành công danh sách sinh viên từ {processed_sheets} sheet.', 'created_count': students_created})
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({'error': str(e), 'traceback': traceback.format_exc()}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        import openpyxl
        from django.http import HttpResponse
        
        # Get query parameters
        faculty = request.query_params.get('faculty')
        class_name = request.query_params.get('className') or request.query_params.get('class_name')
        
        # Filter queryset
        queryset = Student.objects.all()
        if faculty and faculty != 'all':
            queryset = queryset.filter(faculty__iexact=faculty)
        if class_name and class_name != 'all':
            queryset = queryset.filter(class_info__name__iexact=class_name)
            
        # Group by class name
        from collections import defaultdict
        grouped_students = defaultdict(list)
        for student in queryset:
            c_name = student.class_info.name if student.class_info else "Chưa xếp lớp"
            grouped_students[c_name].append(student)
            
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        if not grouped_students:
            # If empty, create at least one empty sheet
            wb.create_sheet(title="Sinh viên")
        else:
            for c_name, students_list in grouped_students.items():
                sheet_title = c_name[:30]
                sheet = wb.create_sheet(title=sheet_title)
                
                headers = ["Mã SV", "Họ và tên", "Lớp", "Khoa", "Giới tính", "Email", "Số điện thoại", "Mật khẩu"]
                sheet.append(headers)
                
                for s in students_list:
                    if not s.user:
                        username = s.student_id.lower()
                        import random
                        if User.objects.filter(username=username).exists():
                            username = f"{username}_{random.randint(1000, 9999)}"
                        
                        random_password = generate_random_password()
                        user_obj = User.objects.create_user(
                            username=username,
                            email=s.email,
                            password=random_password,
                            role='student',
                            full_name=s.full_name,
                            student_id=s.student_id,
                            is_first_login=True,
                            plain_password=random_password
                        )
                        s.user = user_obj
                        s.save()
                    
                    plain_password = s.user.plain_password
                    if not plain_password:
                        random_password = generate_random_password()
                        s.user.set_password(random_password)
                        s.user.plain_password = random_password
                        s.user.save()
                        plain_password = random_password
                        
                    sheet.append([
                        s.student_id,
                        s.full_name,
                        s.class_info.name if s.class_info else "Chưa xếp lớp",
                        s.faculty,
                        s.gender,
                        s.email,
                        s.phone or "",
                        plain_password
                    ])
                
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="danh_sach_sinh_vien.xlsx"'
        wb.save(response)
        return response

# 4. Criteria set and criterion ViewSets
class CriteriaSetViewSet(viewsets.ModelViewSet):
    queryset = CriteriaSet.objects.prefetch_related('criteria').all()
    serializer_class = CriteriaSetSerializer
    permission_classes = [CriteriaAdminOrReadOnly]

    def create(self, request, *args, **kwargs):
        data = request.data.copy()
        activate = bool(data.get('is_active', False))
        clone_from = data.pop('clone_from', None)
        data['is_active'] = False
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)

        with transaction.atomic():
            criteria_set = serializer.save()
            if clone_from:
                source = get_object_or_404(
                    CriteriaSet.objects.prefetch_related('criteria__groups__sub_items'),
                    pk=clone_from
                )
                for source_criterion in source.criteria.all():
                    criterion = Criterion.objects.create(
                        criteria_set=criteria_set,
                        code=source_criterion.code,
                        name=source_criterion.name,
                        max_score=source_criterion.max_score,
                        description=source_criterion.description,
                        is_manual=source_criterion.is_manual,
                    )
                    for source_group in source_criterion.groups.all():
                        group = GroupCriterion.objects.create(
                            criterion=criterion,
                            name=source_group.name,
                            is_single_choice=source_group.is_single_choice,
                        )
                        SubItem.objects.bulk_create([
                            SubItem(group=group, name=item.name, max_score=item.max_score)
                            for item in source_group.sub_items.all()
                        ])
            if activate:
                CriteriaSet.objects.exclude(pk=criteria_set.pk).update(is_active=False)
                criteria_set.is_active = True
                criteria_set.save(update_fields=('is_active', 'updated_at'))

        return Response(self.get_serializer(criteria_set).data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        activate = bool(request.data.get('is_active', False))
        with transaction.atomic():
            response = super().update(request, *args, **kwargs)
            if activate:
                CriteriaSet.objects.exclude(pk=self.get_object().pk).update(is_active=False)
        return response

    def destroy(self, request, *args, **kwargs):
        criteria_set = self.get_object()
        if criteria_set.is_active:
            return Response(
                {'detail': 'Không thể xóa bộ tiêu chí đang được áp dụng.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if criteria_set.evaluations.exists():
            return Response(
                {'detail': 'Không thể xóa bộ tiêu chí đã được dùng trong phiếu đánh giá.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['post'])
    def activate(self, request, pk=None):
        criteria_set = self.get_object()
        if not criteria_set.criteria.exists():
            return Response(
                {'detail': 'Không thể áp dụng một bộ chưa có tiêu chí.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        with transaction.atomic():
            CriteriaSet.objects.exclude(pk=criteria_set.pk).update(is_active=False)
            criteria_set.is_active = True
            criteria_set.save(update_fields=('is_active', 'updated_at'))
        return Response(self.get_serializer(criteria_set).data)


class CriterionViewSet(viewsets.ModelViewSet):
    queryset = Criterion.objects.select_related('criteria_set').prefetch_related('groups__sub_items').all()
    serializer_class = CriterionSerializer
    permission_classes = [CriteriaAdminOrReadOnly]

    def get_queryset(self):
        queryset = self.queryset
        if self.action != 'list':
            return queryset
        criteria_set_id = self.request.query_params.get('criteria_set')
        if criteria_set_id:
            return queryset.filter(criteria_set_id=criteria_set_id)
        if self.request.query_params.get('all') == 'true':
            return queryset
        active_set = CriteriaSet.objects.filter(is_active=True).first()
        return queryset.filter(criteria_set=active_set) if active_set else queryset.none()

    def create(self, request, *args, **kwargs):
        code = request.data.get('code')
        name = request.data.get('name')
        max_score = request.data.get('maxScore') or request.data.get('max_score', 0)
        description = request.data.get('description', '')
        is_manual = request.data.get('isManual', request.data.get('is_manual', False))
        groups_data = request.data.get('groups', [])
        criteria_set_id = request.data.get('criteria_set')
        criteria_set = (
            get_object_or_404(CriteriaSet, pk=criteria_set_id)
            if criteria_set_id else CriteriaSet.objects.filter(is_active=True).first()
        )
        if not criteria_set:
            return Response(
                {'detail': 'Vui lòng tạo hoặc kích hoạt một bộ tiêu chí trước.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        criterion = Criterion.objects.create(
            criteria_set=criteria_set,
            code=code,
            name=name,
            max_score=max_score,
            description=description,
            is_manual=is_manual,
        )

        from .models import GroupCriterion, SubItem
        for g in groups_data:
            group = GroupCriterion.objects.create(
                criterion=criterion,
                name=g.get('name', ''),
                is_single_choice=g.get('isSingleChoice') or g.get('is_single_choice', False)
            )
            for s in g.get('subItems', []):
                SubItem.objects.create(
                    group=group,
                    name=s.get('name', ''),
                    max_score=s.get('maxScore') or s.get('max_score', 0)
                )

        return Response(CriterionSerializer(criterion).data, status=status.HTTP_201_CREATED)

    def update(self, request, pk=None, *args, **kwargs):
        criterion = self.get_object()
        if criterion.criteria_set.evaluations.exists():
            return Response(
                {'detail': 'Bộ tiêu chí đã được sử dụng nên không thể sửa cấu trúc.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        criterion.code = request.data.get('code', criterion.code)
        criterion.name = request.data.get('name', criterion.name)
        criterion.max_score = request.data.get('maxScore') or request.data.get('max_score', criterion.max_score)
        criterion.description = request.data.get('description', criterion.description)
        criterion.is_manual = request.data.get('isManual', request.data.get('is_manual', criterion.is_manual))
        criterion.save()

        groups_data = request.data.get('groups')
        if groups_data is not None:
            criterion.groups.all().delete()
            from .models import GroupCriterion, SubItem
            for g in groups_data:
                group = GroupCriterion.objects.create(
                    criterion=criterion,
                    name=g.get('name', ''),
                    is_single_choice=g.get('isSingleChoice') or g.get('is_single_choice', False)
                )
                for s in g.get('subItems', []):
                    SubItem.objects.create(
                        group=group,
                        name=s.get('name', ''),
                        max_score=s.get('maxScore') or s.get('max_score', 0)
                    )

        return Response(CriterionSerializer(criterion).data)

    def destroy(self, request, *args, **kwargs):
        criterion = self.get_object()
        if criterion.criteria_set.evaluations.exists():
            return Response(
                {'detail': 'Bộ tiêu chí đã được sử dụng nên không thể xóa tiêu chí.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        return super().destroy(request, *args, **kwargs)

def classify_training_score(total_score):
    try:
        config, created = SystemConfig.objects.get_or_create(
            key="training_score_scale",
            defaults={
                "value": [
                    {"min_score": 90, "label": "Xuất sắc"},
                    {"min_score": 80, "label": "Giỏi"},
                    {"min_score": 65, "label": "Khá"},
                    {"min_score": 50, "label": "Trung bình"},
                    {"min_score": 35, "label": "Yếu"},
                    {"min_score": 0, "label": "Kém"}
                ],
                "description": "Thang điểm xếp loại điểm rèn luyện"
            }
        )
        scale = config.value
        scale = sorted(scale, key=lambda x: x.get('min_score', 0), reverse=True)
        for entry in scale:
            if total_score >= entry.get('min_score', 0):
                return entry.get('label', 'Kém')
    except Exception:
        pass
    if total_score >= 90:
        return "Xuất sắc"
    if total_score >= 80:
        return "Giỏi"
    if total_score >= 65:
        return "Khá"
    if total_score >= 50:
        return "Trung bình"
    if total_score >= 35:
        return "Yếu"
    return "Kém"


def sync_evaluation_with_transcript(evaluation):
    """Automatically synchronizes evaluation GPA and classification with the latest imported transcript."""
    from .models import TranscriptImportItem, EvaluationDetail, SubItem
    import re

    def clean_name(name):
        name = name.lower()
        accents = {
            'a': 'áàảãạăắằẳẵặâấầẩẫậ',
            'd': 'đ',
            'e': 'éèẻẽẹêếềểễệ',
            'i': 'íìỉĩị',
            'o': 'óòỏõọôốồổỗộơớờởỡợ',
            'u': 'úùủũụưứừửữự',
            'y': 'ýỳỷỹỵ'
        }
        for char, group in accents.items():
            for g in group:
                name = name.replace(g, char)
        return re.sub(r'[^a-z0-9]', '', name)

    transcript_item = TranscriptImportItem.objects.filter(
        student=evaluation.student,
        transcript_import__semester=evaluation.semester,
        transcript_import__school_year=evaluation.year,
        transcript_import__status='IMPORTED'
    ).order_by('-transcript_import__uploaded_at', '-id').first()

    if transcript_item:
        changed = False
        if evaluation.academic_gpa != transcript_item.gpa:
            evaluation.academic_gpa = transcript_item.gpa
            changed = True
        if evaluation.academic_classification != transcript_item.classification:
            evaluation.academic_classification = transcript_item.classification
            changed = True

        if evaluation.criteria_set:
            academic_criteria = [
                c for c in evaluation.criteria_set.criteria.all()
                if 'hocluc' in clean_name(c.name) or 'hoctap' in clean_name(c.name) or 'academic' in clean_name(c.name)
            ]
            for criterion in academic_criteria:
                if criterion.is_manual:
                    continue
                sub_items = SubItem.objects.filter(group__criterion=criterion)
                classif_clean = clean_name(transcript_item.classification)
                
                matched_sub_item = None
                for sub in sub_items:
                    sub_clean = clean_name(sub.name)
                    if classif_clean in sub_clean or sub_clean in classif_clean:
                        matched_sub_item = sub
                        break
                
                if not matched_sub_item and sub_items.exists():
                    matched_sub_item = sub_items.first()

                if matched_sub_item:
                    score_val = matched_sub_item.max_score
                    deleted_count, _ = EvaluationDetail.objects.filter(
                        evaluation=evaluation,
                        sub_item__group__criterion=criterion
                    ).exclude(sub_item=matched_sub_item).delete()
                    if deleted_count > 0:
                        changed = True
                    
                    detail, det_created = EvaluationDetail.objects.update_or_create(
                        evaluation=evaluation,
                        sub_item=matched_sub_item,
                        defaults={'score': score_val}
                    )
                    if det_created or changed:
                        changed = True

        if changed:
            evaluation.save(update_fields=('academic_gpa', 'academic_classification'))
            recalculate_evaluation_score(evaluation)
            rebalance_training_score(evaluation.student)


def recalculate_evaluation_score(evaluation):
    criteria_set = evaluation.criteria_set
    if not criteria_set:
        evaluation.raw_score = 0
        evaluation.base_score = 0
        evaluation.total_score = 0
        evaluation.classification = classify_training_score(0)
        evaluation.save(update_fields=('raw_score', 'base_score', 'total_score', 'classification'))
        return

    total_score = 0
    for criterion in criteria_set.criteria.all():
        crit_score = sum(
            detail.score
            for detail in EvaluationDetail.objects.filter(
                evaluation=evaluation,
                sub_item__group__criterion=criterion,
            )
        )
        total_score += max(0, min(criterion.max_score, crit_score))

    evaluation.raw_score = total_score
    evaluation.base_score = total_score
    evaluation.total_score = total_score
    evaluation.classification = classify_training_score(total_score)
    evaluation.save()


@transaction.atomic
def rebalance_training_score(student):
    """Move surplus points to the nearest deficient semester.

    A later surplus first repairs previous semesters. If previous semesters
    are already full, it remains available for the next deficient semester.
    """
    semester_order = {'HK1': 1, 'HK2': 2, 'HK3': 3}
    evaluations = list(
        Evaluation.objects.filter(student=student).select_related('criteria_set')
    )
    evaluations.sort(
        key=lambda item: (
            item.year.split('-')[0] if item.year else '',
            semester_order.get(item.semester, 99),
            item.id,
        )
    )

    states = []
    for evaluation in evaluations:
        maximum = sum(
            criterion.max_score
            for criterion in evaluation.criteria_set.criteria.all()
        ) if evaluation.criteria_set else 100
        base_score = max(0, min(maximum, evaluation.raw_score))
        states.append({
            'evaluation': evaluation,
            'maximum': maximum,
            'base': base_score,
            'deficit': max(0, maximum - base_score),
            'surplus': max(0, evaluation.raw_score - maximum),
            'carry_in': 0,
            'carry_out': 0,
        })

    for source_index, source in enumerate(states):
        available = source['surplus']
        if available <= 0:
            continue

        # Newer surplus repairs the closest previous deficient semester first.
        target_indexes = list(range(source_index - 1, -1, -1))
        # Any remaining balance is reserved for the following semesters.
        target_indexes += list(range(source_index + 1, len(states)))
        for target_index in target_indexes:
            target = states[target_index]
            if target['deficit'] <= 0:
                continue
            transferred = min(available, target['deficit'])
            target['deficit'] -= transferred
            target['carry_in'] += transferred
            source['carry_out'] += transferred
            available -= transferred
            if available <= 0:
                break
        source['surplus_balance'] = available

    for state in states:
        evaluation = state['evaluation']
        evaluation.base_score = state['base']
        evaluation.carry_in = state['carry_in']
        evaluation.carry_out = state['carry_out']
        evaluation.surplus_balance = state.get('surplus_balance', 0)
        evaluation.total_score = min(
            state['maximum'],
            state['base'] + state['carry_in'],
        )
        evaluation.classification = classify_training_score(evaluation.total_score)
        evaluation.save(update_fields=(
            'base_score', 'carry_in', 'carry_out', 'surplus_balance',
            'total_score', 'classification',
        ))


# 5. Evaluation ViewSet
class EvaluationViewSet(viewsets.ModelViewSet):
    queryset = Evaluation.objects.all()
    serializer_class = EvaluationSerializer
    permission_classes = [permissions.AllowAny]

    def get_queryset(self):
        user = self.request.user
        queryset = Evaluation.objects.select_related(
            'student__class_info',
            'criteria_set',
        ).prefetch_related(
            'criteria_set__criteria',
            'details__sub_item__group__criterion',
        )
        
        if user.is_authenticated:
            if 'class_monitor' in user.roles:
                student = getattr(user, 'student_profile', None)
                if student and student.class_info:
                    queryset = queryset.filter(student__class_info=student.class_info)
                else:
                    queryset = queryset.filter(student__user=user)
            elif user.role == 'student':
                queryset = queryset.filter(student__user=user).exclude(status='draft')
            elif user.role == 'advisor':
                queryset = queryset.filter(student__class_info__advisor=user)
            
        class_name = self.request.query_params.get('className') or self.request.query_params.get('class_name')
        semester = self.request.query_params.get('semester')
        year = self.request.query_params.get('year')
        status_param = self.request.query_params.get('status')
        
        student_id = self.request.query_params.get('studentId') or self.request.query_params.get('student_id')
        
        if class_name:
            queryset = queryset.filter(student__class_info__name=class_name)
        if semester:
            queryset = queryset.filter(semester=semester)
        if year:
            queryset = queryset.filter(year=year)
        if status_param:
            queryset = queryset.filter(status=status_param)
        if student_id:
            queryset = queryset.filter(student__student_id=student_id)
            
        return queryset

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        evals_list = list(queryset)
        if len(evals_list) <= 120:
            for ev in evals_list:
                sync_evaluation_with_transcript(ev)
        serializer = self.get_serializer(evals_list, many=True)
        return Response(serializer.data)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        sync_evaluation_with_transcript(instance)
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    def _error_res(self, code, message, details=None, status_code=400):
        return Response({
            'success': False,
            'code': code,
            'message': message,
            'details': details or {}
        }, status=status_code)

    def update(self, request, *args, **kwargs):
        evaluation = self.get_object()
        client_version = request.data.get('clientVersion')
        
        from .services.workflow_guard import validate_evaluation_write_access, VersionConflictException, WorkflowLockedException, IllegalUpdateException
        from django.core.exceptions import ValidationError
        
        try:
            validate_evaluation_write_access(
                evaluation=evaluation,
                user=request.user,
                client_version=client_version,
                request=request
            )
        except VersionConflictException as e:
            return self._error_res('VERSION_CONFLICT', str(e), {'serverVersion': e.server_version}, status.HTTP_409_CONFLICT)
        except WorkflowLockedException as e:
            return self._error_res('WORKFLOW_LOCKED', str(e), {}, status.HTTP_403_FORBIDDEN)
        except IllegalUpdateException as e:
            return self._error_res('UNAUTHORIZED', str(e), {}, status.HTTP_403_FORBIDDEN)
        except ValidationError as e:
            return self._error_res('VALIDATION_ERROR', str(e), {}, status.HTTP_400_BAD_REQUEST)
            
        response = super().update(request, *args, **kwargs)
        if response.status_code == status.HTTP_200_OK:
            evaluation.refresh_from_db()
            evaluation.version += 1
            evaluation.save(update_fields=['version'])
            response.data['serverVersion'] = evaluation.version
            # Log the edit
            from .services.workflow_guard import log_audit
            log_audit(user=request.user, action='Draft Updated', entity_id=evaluation.id, before_value='', after_value=f"version={evaluation.version}", request=request)
        return response

    def partial_update(self, request, *args, **kwargs):
        evaluation = self.get_object()
        client_version = request.data.get('clientVersion')
        
        from .services.workflow_guard import validate_evaluation_write_access, VersionConflictException, WorkflowLockedException, IllegalUpdateException
        from django.core.exceptions import ValidationError
        
        try:
            validate_evaluation_write_access(
                evaluation=evaluation,
                user=request.user,
                client_version=client_version,
                request=request
            )
        except VersionConflictException as e:
            return self._error_res('VERSION_CONFLICT', str(e), {'serverVersion': e.server_version}, status.HTTP_409_CONFLICT)
        except WorkflowLockedException as e:
            return self._error_res('WORKFLOW_LOCKED', str(e), {}, status.HTTP_403_FORBIDDEN)
        except IllegalUpdateException as e:
            return self._error_res('UNAUTHORIZED', str(e), {}, status.HTTP_403_FORBIDDEN)
        except ValidationError as e:
            return self._error_res('VALIDATION_ERROR', str(e), {}, status.HTTP_400_BAD_REQUEST)
            
        response = super().partial_update(request, *args, **kwargs)
        if response.status_code == status.HTTP_200_OK:
            evaluation.refresh_from_db()
            evaluation.version += 1
            evaluation.save(update_fields=['version'])
            response.data['serverVersion'] = evaluation.version
            from .services.workflow_guard import log_audit
            log_audit(user=request.user, action='Draft Updated', entity_id=evaluation.id, before_value='', after_value=f"version={evaluation.version}", request=request)
        return response

    def create(self, request, *args, **kwargs):
        import time
        from django.db import OperationalError

        for attempt in range(5):
            try:
                with transaction.atomic():
                    student_id = request.data.get('studentId') or request.data.get('student_id')
                    student = get_object_or_404(Student, student_id=student_id)
                    semester = request.data.get('semester')
                    year = request.data.get('year')
                    scores_data = request.data.get('scores', {}) # dict of subitem_id -> score
                    note = request.data.get('note', '')
                    academic_gpa = request.data.get('academicGpa', request.data.get('academic_gpa'))
                    academic_classification = request.data.get(
                        'academicClassification',
                        request.data.get('academic_classification', '')
                    )
                    requested_raw_score = request.data.get('rawScore', request.data.get('raw_score'))
                    status_param = request.data.get('status', 'draft')
                    requested_set_id = request.data.get('criteriaSet') or request.data.get('criteria_set')

                    existing_evaluation = Evaluation.objects.filter(
                        student=student, semester=semester, year=year
                    ).first()

                    if requested_set_id:
                        criteria_set = get_object_or_404(CriteriaSet, pk=requested_set_id)
                    else:
                        active_set = (
                            CriteriaSet.objects.filter(
                                semester=semester, academic_year=year, is_active=True
                            ).first()
                            or CriteriaSet.objects.filter(is_active=True).first()
                        )
                        if active_set:
                            criteria_set = active_set
                        elif existing_evaluation and existing_evaluation.criteria_set:
                            criteria_set = existing_evaluation.criteria_set
                        else:
                            criteria_set = None
                    if not criteria_set:
                        return Response(
                            {'detail': 'Chưa có bộ tiêu chí nào được quản trị viên kích hoạt.'},
                            status=status.HTTP_400_BAD_REQUEST
                        )

                    # Create or update evaluation
                    evaluation, created = Evaluation.objects.update_or_create(
                        student=student, semester=semester, year=year,
                        defaults={
                            'note': note,
                            'academic_gpa': academic_gpa if academic_gpa not in ('', None) else None,
                            'academic_classification': academic_classification or '',
                            'status': status_param,
                            'class_confirmed': False,
                            'criteria_set': criteria_set,
                        }
                    )

                    # Clear old details
                    evaluation.details.all().delete()

                    # Write new details and calculate total score
                    from .models import SubItem
                    for sub_item_id, score_val in scores_data.items():
                        try:
                            sub_item = SubItem.objects.get(
                                id=sub_item_id,
                                group__criterion__criteria_set=criteria_set
                            )
                            EvaluationDetail.objects.create(
                                evaluation=evaluation,
                                sub_item=sub_item,
                                score=score_val
                            )
                        except SubItem.DoesNotExist:
                            pass

                    sync_evaluation_with_transcript(evaluation)
                    recalculate_evaluation_score(evaluation)
                    if requested_raw_score not in (None, ''):
                        requested_total = int(float(requested_raw_score))
                        if requested_total > evaluation.raw_score:
                            evaluation.raw_score = requested_total
                            evaluation.save(update_fields=('raw_score',))
                    rebalance_training_score(student)
                    evaluation.refresh_from_db()
                    return Response(EvaluationSerializer(evaluation).data, status=status.HTTP_201_CREATED)
            except OperationalError as e:
                if 'locked' in str(e).lower() and attempt < 4:
                    time.sleep(0.1 * (attempt + 1))
                    continue
                raise

    @action(detail=False, methods=['post'], url_path='bulk-init')
    def bulk_init(self, request):
        payload_data = request.data
        if not isinstance(payload_data, list):
            return Response({'detail': 'Payload must be a list of evaluations.'}, status=status.HTTP_400_BAD_REQUEST)
        
        from .services.evaluation_bulk_service import launch_bulk_init_job
        job = launch_bulk_init_job(payload_data)
        
        return Response({
            'jobId': job.id,
            'status': job.status,
            'progress': job.progress,
            'total': job.total
        }, status=status.HTTP_202_ACCEPTED)

    @action(detail=False, methods=['post'], url_path='bulk-update-status')
    def bulk_update_status(self, request):
        status_param = request.data.get('status')
        evaluation_ids = request.data.get('evaluationIds')
        criteria_set_id = request.data.get('criteriaSetId')
        semester = request.data.get('semester')
        year = request.data.get('year')
        class_name = request.data.get('className')

        if status_param not in ['draft', 'published', 'class_pending', 'advisor_pending', 'pending', 'approved', 'rejected', 'locked']:
            return Response({'detail': 'Trạng thái không hợp lệ.'}, status=status.HTTP_400_BAD_REQUEST)

        queryset = Evaluation.objects.all()
        if evaluation_ids:
            queryset = queryset.filter(id__in=evaluation_ids)
        else:
            if criteria_set_id:
                queryset = queryset.filter(criteria_set_id=criteria_set_id)
            if semester:
                queryset = queryset.filter(semester=semester)
            if year:
                queryset = queryset.filter(year=year)
            if class_name:
                queryset = queryset.filter(student__class_info__name=class_name)

        count = queryset.update(status=status_param)
        return Response({'success': True, 'updated_count': count})

    @action(detail=False, methods=['get'], url_path='bulk-job/(?P<job_id>[^/.]+)')
    def bulk_job_status(self, request, job_id=None):
        from .models import EvaluationJob
        job = get_object_or_404(EvaluationJob, pk=job_id)
        return Response({
            'jobId': job.id,
            'status': job.status,
            'progress': job.progress,
            'total': job.total,
            'errorMessage': job.error_message
        })

    @action(detail=True, methods=['post'], url_path='submit-self-assessment')
    def submit_self_assessment(self, request, pk=None):
        evaluation = self.get_object()
        client_version = request.data.get('clientVersion')
        
        from .services.workflow_guard import validate_evaluation_write_access, VersionConflictException, WorkflowLockedException, IllegalUpdateException
        from django.core.exceptions import ValidationError
        
        try:
            validate_evaluation_write_access(
                evaluation=evaluation,
                user=request.user,
                client_version=client_version,
                request=request
            )
        except VersionConflictException as e:
            return self._error_res('VERSION_CONFLICT', str(e), {'serverVersion': e.server_version}, status.HTTP_409_CONFLICT)
        except WorkflowLockedException as e:
            return self._error_res('WORKFLOW_LOCKED', str(e), {}, status.HTTP_403_FORBIDDEN)
        except IllegalUpdateException as e:
            return self._error_res('UNAUTHORIZED', str(e), {}, status.HTTP_403_FORBIDDEN)
        except ValidationError as e:
            return self._error_res('VALIDATION_ERROR', str(e), {}, status.HTTP_400_BAD_REQUEST)

        scores_data = request.data.get('scores', {})
        note = request.data.get('note', evaluation.note or '')
        manual_subitems = SubItem.objects.filter(
            group__criterion__criteria_set=evaluation.criteria_set,
            group__criterion__is_manual=True,
        )
        manual_subitem_ids = set(manual_subitems.values_list('id', flat=True))

        with transaction.atomic():
            if manual_subitem_ids:
                evaluation.details.filter(sub_item_id__in=manual_subitem_ids).delete()
            for sub_item_id, score_val in scores_data.items():
                try:
                    sub_item_id_int = int(sub_item_id)
                    if sub_item_id_int not in manual_subitem_ids:
                        continue
                    sub_item = manual_subitems.get(id=sub_item_id_int)
                    score = int(float(score_val or 0))
                except (ValueError, TypeError, SubItem.DoesNotExist):
                    continue
                EvaluationDetail.objects.update_or_create(
                    evaluation=evaluation,
                    sub_item=sub_item,
                    defaults={'score': score},
                )

            evaluation.note = note
            evaluation.status = 'class_pending'
            evaluation.class_confirmed = False
            evaluation.self_submitted_at = timezone.now()
            evaluation.version += 1
            evaluation.save(update_fields=('note', 'status', 'class_confirmed', 'self_submitted_at', 'version'))
            recalculate_evaluation_score(evaluation)

            # Send notifications
            if evaluation.student.class_info:
                monitors = User.objects.filter(
                    student_id__in=StudentClassPosition.objects.filter(
                        class_info=evaluation.student.class_info,
                        position__name__in=['Lớp trưởng', 'Lớp phó']
                    ).values_list('student__student_id', flat=True)
                )
                for monitor in monitors:
                    create_notification(
                        user=monitor,
                        title="Phiếu tự đánh giá mới cần duyệt",
                        message=f"Sinh viên {evaluation.student.full_name} ({evaluation.student.student_id}) lớp {evaluation.student.class_info.name} đã nộp phiếu tự đánh giá HK{evaluation.semester} {evaluation.year}.",
                        type='evaluation',
                        level='warning',
                        action_url='/class-review'
                    )
                if evaluation.student.class_info.advisor:
                    create_notification(
                        user=evaluation.student.class_info.advisor,
                        title="Phiếu tự đánh giá mới cần duyệt",
                        message=f"Sinh viên {evaluation.student.full_name} ({evaluation.student.student_id}) lớp {evaluation.student.class_info.name} đã nộp phiếu tự đánh giá HK{evaluation.semester} {evaluation.year}.",
                        type='evaluation',
                        level='warning',
                        action_url='/approvals'
                    )

        rebalance_training_score(evaluation.student)
        evaluation.refresh_from_db()
        
        serializer_data = EvaluationSerializer(evaluation).data
        serializer_data['serverVersion'] = evaluation.version
        return Response(serializer_data)

    @action(detail=True, methods=['post'], url_path='save-draft')
    def save_draft(self, request, pk=None):
        import time
        start_time = time.time()
        evaluation = self.get_object()
        scores_data = request.data.get('scores', {})
        note = request.data.get('note')
        client_version = request.data.get('clientVersion')

        from .services.workflow_guard import VersionConflictException, log_audit
        from .services.evaluation_bulk_service import save_evaluation_draft_bulk
        from django.core.exceptions import ValidationError
        from .services.evaluation_monitor import record_autosave_metrics

        try:
            res_data = save_evaluation_draft_bulk(
                evaluation=evaluation,
                scores_data=scores_data,
                note=note,
                user=request.user,
                client_version=client_version,
                request=request,
                recalculate_fn=recalculate_evaluation_score,
                rebalance_fn=rebalance_training_score
            )
            duration_ms = int((time.time() - start_time) * 1000)
            record_autosave_metrics(success=True, is_conflict=False, duration_ms=duration_ms)
            
            # Log standard audit event
            log_audit(user=request.user, action='Auto Save', entity_id=evaluation.id, before_value='', after_value=f"version={evaluation.version}", request=request)
            return Response(res_data, status=status.HTTP_200_OK)
        except VersionConflictException as e:
            record_autosave_metrics(success=False, is_conflict=True)
            return self._error_res('VERSION_CONFLICT', str(e), {'serverVersion': e.server_version}, status.HTTP_409_CONFLICT)
        except ValidationError as e:
            record_autosave_metrics(success=False, is_conflict=False)
            err_msg = e.message if hasattr(e, 'message') else str(e)
            code = 'SESSION_INVALID' if 'phiên hoạt động' in err_msg else 'VALIDATION_ERROR'
            status_code = status.HTTP_403_FORBIDDEN if code == 'SESSION_INVALID' else status.HTTP_400_BAD_REQUEST
            return self._error_res(code, err_msg, {}, status_code)
        except Exception as e:
            record_autosave_metrics(success=False, is_conflict=False)
            return self._error_res('SERVER_ERROR', str(e), {}, status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'], url_path='review')
    def review(self, request, pk=None):
        evaluation = self.get_object()
        client_version = request.data.get('clientVersion')
        
        from .services.workflow_guard import validate_evaluation_write_access, VersionConflictException, WorkflowLockedException, IllegalUpdateException, log_audit
        from django.core.exceptions import ValidationError
        
        try:
            validate_evaluation_write_access(
                evaluation=evaluation,
                user=request.user,
                client_version=client_version,
                request=request
            )
        except VersionConflictException as e:
            return self._error_res('VERSION_CONFLICT', str(e), {'serverVersion': e.server_version}, status.HTTP_409_CONFLICT)
        except WorkflowLockedException as e:
            return self._error_res('WORKFLOW_LOCKED', str(e), {}, status.HTTP_403_FORBIDDEN)
        except IllegalUpdateException as e:
            return self._error_res('UNAUTHORIZED', str(e), {}, status.HTTP_403_FORBIDDEN)
        except ValidationError as e:
            return self._error_res('VALIDATION_ERROR', str(e), {}, status.HTTP_400_BAD_REQUEST)

        review_note = request.data.get('reviewNote', '')
        evaluation.class_confirmed = True
        evaluation.status = 'advisor_pending'
        evaluation.review_note = review_note
        evaluation.version += 1
        evaluation.save()

        # Log standard audit event
        log_audit(user=request.user, action='Draft Updated', entity_id=evaluation.id, before_value='', after_value=f"status=class_reviewed_version={evaluation.version}", request=request)

        # Send notifications
        student_user = User.objects.filter(student_id=evaluation.student.student_id).first()
        if student_user:
            create_notification(
                user=student_user,
                title="Phiếu rèn luyện được duyệt ở cấp lớp",
                message=f"Phiếu rèn luyện HK{evaluation.semester} {evaluation.year} của bạn đã được ban cán sự lớp thông qua và chuyển lên Cố vấn học tập.",
                type='evaluation',
                level='success',
                action_url='/'
            )
        if evaluation.student.class_info and evaluation.student.class_info.advisor:
            create_notification(
                user=evaluation.student.class_info.advisor,
                title="Phiếu rèn luyện mới chờ phê duyệt",
                message=f"Phiếu rèn luyện của sinh viên {evaluation.student.full_name} ({evaluation.student.student_id}) đã được lớp thông qua và chờ bạn phê duyệt.",
                type='evaluation',
                level='warning',
                action_url='/approvals'
            )

        serializer_data = EvaluationSerializer(evaluation).data
        serializer_data['serverVersion'] = evaluation.version
        return Response(serializer_data)

    @action(detail=True, methods=['post'], url_path='approve')
    def approve(self, request, pk=None):
        evaluation = self.get_object()
        status_param = request.data.get('status')
        review_note = request.data.get('reviewNote', '')
        client_version = request.data.get('clientVersion')
        
        from .services.workflow_guard import validate_evaluation_write_access, VersionConflictException, WorkflowLockedException, IllegalUpdateException, log_audit
        from django.core.exceptions import ValidationError
        
        try:
            validate_evaluation_write_access(
                evaluation=evaluation,
                user=request.user,
                client_version=client_version,
                request=request
            )
        except VersionConflictException as e:
            return self._error_res('VERSION_CONFLICT', str(e), {'serverVersion': e.server_version}, status.HTTP_409_CONFLICT)
        except WorkflowLockedException as e:
            return self._error_res('WORKFLOW_LOCKED', str(e), {}, status.HTTP_403_FORBIDDEN)
        except IllegalUpdateException as e:
            return self._error_res('UNAUTHORIZED', str(e), {}, status.HTTP_403_FORBIDDEN)
        except ValidationError as e:
            return self._error_res('VALIDATION_ERROR', str(e), {}, status.HTTP_400_BAD_REQUEST)
        
        if status_param not in ['approved', 'rejected', 'pending', 'published', 'locked']:
            return self._error_res('VALIDATION_ERROR', 'Trạng thái phê duyệt không hợp lệ.', {}, status.HTTP_400_BAD_REQUEST)
            
        evaluation.status = status_param
        evaluation.review_note = review_note
        evaluation.version += 1
        evaluation.save()

        # Audit standard action
        audit_action = f'Status Updated to {status_param.capitalize()}'
        if status_param == 'approved':
            audit_action = 'Approved'
        elif status_param == 'rejected':
            audit_action = 'Rejected'
        elif status_param == 'published':
            audit_action = 'Published'
        elif status_param == 'locked':
            audit_action = 'Locked'
            
        log_audit(user=request.user, action=audit_action, entity_id=evaluation.id, before_value='', after_value=f"status={status_param}_version={evaluation.version}", request=request)

        # Send notifications
        student_user = User.objects.filter(student_id=evaluation.student.student_id).first()
        if student_user:
            if status_param == 'approved':
                create_notification(
                    user=student_user,
                    title="Phiếu rèn luyện đã được duyệt hoàn tất",
                    message=f"Cố vấn học tập đã duyệt hoàn tất phiếu rèn luyện HK{evaluation.semester} {evaluation.year} của bạn với tổng số điểm là {evaluation.total_score}.",
                    type='evaluation',
                    level='success',
                    action_url='/'
                )
                if evaluation.student.email:
                    try:
                        from django.core.mail import send_mail
                        from django.conf import settings
                        email_subject = f"[ITC Point] Kết quả điểm rèn luyện chính thức HK{evaluation.semester} {evaluation.year}"
                        email_message = f"Chào {evaluation.student.full_name},\n\nĐiểm rèn luyện chính thức của bạn trong HK{evaluation.semester} {evaluation.year} đã được phê duyệt hoàn tất.\n\nTổng điểm: {evaluation.total_score}\nXếp loại: {evaluation.classification or 'Chưa xếp loại'}"
                        email_html = f"""
<div style="font-family: 'Segoe UI', Roboto, Helvetica, Arial, sans-serif; max-width: 600px; margin: 0 auto; border: 1px solid #e2e8f0; border-radius: 16px; overflow: hidden; box-shadow: 0 10px 25px -5px rgba(0,0,0,0.05); background-color: #ffffff;">
  <div style="background: linear-gradient(135deg, #10b981, #059669); padding: 32px 24px; text-align: center; color: white;">
    <h1 style="margin: 0; font-size: 24px; font-weight: 800;">Kết Quả Điểm Rèn Luyện</h1>
    <p style="margin: 4px 0 0 0; font-size: 13px; color: rgba(255,255,255,0.85); font-weight: 500;">Phê duyệt hoàn tất cấp Cố vấn học tập</p>
  </div>
  <div style="padding: 40px 32px; background-color: #ffffff;">
    <h2 style="margin-top: 0; color: #1e293b; font-size: 20px; font-weight: 700;">Chào {evaluation.student.full_name},</h2>
    <p style="color: #475569; line-height: 1.6; font-size: 15px;">Hệ thống thông báo: Điểm rèn luyện của bạn trong học kỳ này đã được xét duyệt chính thức với thông tin xếp loại như sau:</p>
    
    <div style="background-color: #f8fafc; border-radius: 12px; padding: 20px; margin: 28px 0; border: 1px solid #f1f5f9; border-left: 4px solid #10b981;">
      <table style="width: 100%; border-collapse: collapse;">
        <tr>
          <td style="padding: 8px 0; font-size: 14px; color: #64748b; font-weight: 600; width: 180px; text-transform: uppercase;">Học kỳ / Năm học:</td>
          <td style="padding: 8px 0; font-size: 15px; color: #0f172a; font-weight: 700;">HK{evaluation.semester} - Năm học {evaluation.year}</td>
        </tr>
        <tr>
          <td style="padding: 8px 0; font-size: 14px; color: #64748b; font-weight: 600; text-transform: uppercase;">Tổng điểm chính thức:</td>
          <td style="padding: 8px 0; font-size: 18px; color: #10b981; font-weight: 800;">{evaluation.total_score} điểm</td>
        </tr>
        <tr>
          <td style="padding: 8px 0; font-size: 14px; color: #64748b; font-weight: 600; text-transform: uppercase;">Xếp loại rèn luyện:</td>
          <td style="padding: 8px 0; font-size: 16px; color: #0f172a; font-weight: 700;">{evaluation.classification or 'Chưa xếp loại'}</td>
        </tr>
      </table>
    </div>

    <div style="text-align: center; margin: 36px 0 28px 0;">
      <a href="http://localhost:8080/" style="background: linear-gradient(135deg, #10b981, #059669); color: #ffffff; padding: 14px 32px; text-decoration: none; border-radius: 10px; font-weight: 700; font-size: 15px; display: inline-block; box-shadow: 0 10px 15px -3px rgba(16, 185, 129, 0.3);">Xem chi tiết bảng điểm</a>
    </div>
  </div>
  <div style="background-color: #f8fafc; padding: 24px; text-align: center; border-top: 1px solid #f1f5f9; color: #94a3b8; font-size: 12px;">
    <p style="margin: 0 0 6px 0; font-weight: 500;">Email này được hệ thống ITC Point gửi tự động.</p>
    <p style="margin: 0;">© 2026 ITC Point. All rights reserved.</p>
  </div>
</div>
"""
                        send_mail(
                            email_subject,
                            email_message,
                            getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@itcpoint.com'),
                            [evaluation.student.email],
                            fail_silently=True,
                            html_message=email_html
                        )
                    except Exception as e:
                        print(f"Error sending evaluation approved email to {evaluation.student.student_id}: {e}")
            elif status_param == 'rejected':
                create_notification(
                    user=student_user,
                    title="Phiếu rèn luyện bị trả lại",
                    message=f"Phiếu rèn luyện HK{evaluation.semester} {evaluation.year} của bạn đã bị từ chối/yêu cầu bổ sung bởi Cố vấn học tập. Lý do: {review_note}",
                    type='evaluation',
                    level='danger',
                    action_url='/'
                )

        serializer_data = EvaluationSerializer(evaluation).data
        serializer_data['serverVersion'] = evaluation.version
        return Response(serializer_data)

    @action(detail=True, methods=['post'], url_path='reject-details')
    def reject_details(self, request, pk=None):
        evaluation = self.get_object()
        rejected_items = request.data.get('rejected_items', [])  # list of {sub_item_id: int, reason: str}
        review_note = request.data.get('reviewNote', '')

        from .services.workflow_guard import validate_evaluation_write_access, log_audit

        try:
            validate_evaluation_write_access(
                evaluation=evaluation,
                user=request.user,
                client_version=request.data.get('clientVersion'),
                request=request
            )
        except Exception as e:
            roles = set(request.user.roles if hasattr(request.user, 'roles') else [request.user.role])
            is_reviewer = bool(roles.intersection({'class_monitor', 'advisor', 'student_affairs', 'admin'}))
            if not is_reviewer:
                return self._error_res('UNAUTHORIZED', str(e), {}, status.HTTP_403_FORBIDDEN)

        with transaction.atomic():
            # Reset any old rejection status for this evaluation's details
            evaluation.details.all().update(is_rejected=False, reject_reason=None)

            # Apply new rejections
            for item in rejected_items:
                sub_id = item.get('sub_item_id')
                reason = item.get('reason', '')
                evaluation.details.filter(sub_item_id=sub_id).update(is_rejected=True, reject_reason=reason)

            evaluation.status = 'rejected'
            evaluation.review_note = review_note or 'Có tiêu chí bị ban cán sự / cố vấn từ chối.'
            evaluation.version += 1
            evaluation.save()

        log_audit(user=request.user, action='Rejected Details', entity_id=evaluation.id, before_value='', after_value=f"status=rejected_version={evaluation.version}", request=request)

        # Send notification to student
        student_user = User.objects.filter(student_id=evaluation.student.student_id).first()
        if student_user:
            create_notification(
                user=student_user,
                title="Phiếu rèn luyện có tiêu chí bị từ chối",
                message=f"Phiếu rèn luyện HK{evaluation.semester} {evaluation.year} của bạn bị từ chối/yêu cầu điều chỉnh một số tiêu chí. Vui lòng kiểm tra lại.",
                type='evaluation',
                level='danger',
                action_url='/'
            )

        serializer_data = EvaluationSerializer(evaluation).data
        serializer_data['serverVersion'] = evaluation.version
        return Response(serializer_data)

    @action(detail=False, methods=['get'], url_path='health')
    def health(self, request):
        if not request.user.is_authenticated or request.user.role != 'admin':
            return self._error_res('UNAUTHORIZED', 'Chỉ Admin mới có quyền truy cập thông tin sức khỏe hệ thống.', {}, status.HTTP_403_FORBIDDEN)
            
        from .services.evaluation_monitor import get_health_metrics
        return Response(get_health_metrics(), status=status.HTTP_200_OK)

# 6. Activity ViewSet
class ActivityViewSet(viewsets.ModelViewSet):
    queryset = Activity.objects.all()
    serializer_class = ActivitySerializer
    permission_classes = [permissions.AllowAny]

    def get_queryset(self):
        queryset = Activity.objects.all()
        is_external_param = self.request.query_params.get('is_external')
        if is_external_param is not None:
            is_external_bool = is_external_param.lower() == 'true'
            queryset = queryset.filter(is_external=is_external_bool)
        return queryset

    def perform_update(self, serializer):
        old_instance = self.get_object()
        old_date = old_instance.date
        old_start_time = old_instance.start_time
        
        instance = serializer.save()
        
        if old_date != instance.date or old_start_time != instance.start_time:
            from django.core.mail import send_mail
            from django.conf import settings
            from django.utils import timezone
            from drl_app.models import ActivityParticipant, User
            
            participants = ActivityParticipant.objects.filter(activity=instance, status='registered')
            
            for p in participants:
                student = p.student
                student_user = User.objects.filter(student_id=student.student_id).first()
                
                if student_user:
                    try:
                        create_notification(
                            user=student_user,
                            title="Thay đổi lịch hoạt động",
                            message=f"Hoạt động '{instance.title}' đã thay đổi thời gian diễn ra.",
                            type='activity',
                            level='warning',
                            action_url=f"/activities/{instance.id}"
                        )
                    except Exception as notif_err:
                        print(f"Error creating reschedule notification: {notif_err}")
                
                if student.email:
                    try:
                        old_time_str = f"{old_start_time.strftime('%H:%M') if old_start_time else '00:00'} ngày {old_date.strftime('%d/%m/%Y') if old_date else 'chưa xác định'}"
                        new_time_str = f"{instance.start_time.strftime('%H:%M') if instance.start_time else '00:00'} ngày {instance.date.strftime('%d/%m/%Y') if instance.date else 'chưa xác định'}"
                        
                        email_subject = f"[ITC Point] Thay đổi lịch hoạt động: {instance.title}"
                        email_message = f"Chào {student.full_name},\n\nBan tổ chức xin thông báo hoạt động '{instance.title}' đã thay đổi thời gian diễn ra từ {old_time_str} sang {new_time_str}.\n\nVui lòng xem thông tin chi tiết trên hệ thống."
                        
                        email_html = f"""
<div style="font-family: 'Segoe UI', Roboto, Helvetica, Arial, sans-serif; max-width: 600px; margin: 0 auto; border: 1px solid #e2e8f0; border-radius: 16px; overflow: hidden; box-shadow: 0 10px 25px -5px rgba(0,0,0,0.05), 0 8px 10px -6px rgba(0,0,0,0.05); background-color: #ffffff;">
  <div style="background: linear-gradient(135deg, #ef4444, #dc2626); padding: 32px 24px; text-align: center; color: white;">
    <h1 style="margin: 0; font-size: 24px; font-weight: 800; letter-spacing: -0.5px;">Thay Đổi Lịch Hoạt Động</h1>
    <p style="margin: 4px 0 0 0; font-size: 13px; color: rgba(255,255,255,0.85); font-weight: 500;">Cập nhật lịch tổ chức hoạt động</p>
  </div>
  <div style="padding: 40px 32px; background-color: #ffffff;">
    <h2 style="margin-top: 0; color: #1e293b; font-size: 20px; font-weight: 700; letter-spacing: -0.5px;">Chào {student.full_name},</h2>
    <p style="color: #475569; line-height: 1.6; font-size: 15px; margin-top: 12px;">Ban tổ chức xin thông báo về việc thay đổi thời gian diễn ra của hoạt động bạn đã đăng ký như sau:</p>
    
    <div style="background-color: #f8fafc; border-radius: 12px; padding: 20px; margin: 28px 0; border: 1px solid #f1f5f9; border-left: 4px solid #ef4444;">
      <table style="width: 100%; border-collapse: collapse;">
        <tr>
          <td style="padding: 8px 0; font-size: 14px; color: #64748b; font-weight: 600; width: 130px; text-transform: uppercase; letter-spacing: 0.5px;">Hoạt động:</td>
          <td style="padding: 8px 0; font-size: 15px; color: #0f172a; font-weight: 700;">{instance.title}</td>
        </tr>
        <tr>
          <td style="padding: 8px 0; font-size: 14px; color: #64748b; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px;">Lịch cũ:</td>
          <td style="padding: 8px 0; font-size: 15px; color: #ef4444; font-weight: 500; text-decoration: line-through;">{old_time_str}</td>
        </tr>
        <tr>
          <td style="padding: 8px 0; font-size: 14px; color: #64748b; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px;">Lịch mới:</td>
          <td style="padding: 8px 0; font-size: 15px; color: #10b981; font-weight: 700;">{new_time_str}</td>
        </tr>
        <tr>
          <td style="padding: 8px 0; font-size: 14px; color: #64748b; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px;">Lý do dời:</td>
          <td style="padding: 8px 0; font-size: 15px; color: #0f172a; font-weight: 500; line-height: 1.4;">Thời gian tổ chức được điều chỉnh để chuẩn bị tốt hơn.</td>
        </tr>
      </table>
    </div>

    <p style="color: #64748b; font-size: 13px; line-height: 1.5; margin: 0;">
      * Mọi quyền lợi đăng ký và điểm danh của hoạt động này sẽ được giữ nguyên theo lịch thi đấu/học tập mới.
    </p>
  </div>
  <div style="background-color: #f8fafc; padding: 24px; text-align: center; border-top: 1px solid #f1f5f9; color: #94a3b8; font-size: 12px; line-height: 1.5;">
    <p style="margin: 0 0 6px 0; font-weight: 500;">Email này được hệ thống ITC Point gửi tự động.</p>
    <p style="margin: 0;">© 2026 ITC Point. All rights reserved.</p>
  </div>
</div>
"""
                        send_mail(
                            email_subject,
                            email_message,
                            getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@itcpoint.com'),
                            [student.email],
                            fail_silently=False,
                            html_message=email_html
                        )
                    except Exception as email_err:
                        print(f"Error sending reschedule email to {student.student_id}: {email_err}")

    @action(
        detail=True,
        methods=['get'],
        url_path='export-participants',
        permission_classes=[permissions.IsAuthenticated],
    )
    def export_participants(self, request, pk=None):
        import openpyxl
        from django.http import HttpResponse
        from openpyxl.styles import Alignment, Font, PatternFill

        activity = self.get_object()
        if not _can_review_activity_attendance(request.user, activity):
            return Response(
                {'error': 'Bạn không có quyền xuất danh sách tham gia hoạt động này.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        participants = list(
            activity.participants.select_related('student').order_by(
                'student__student_id',
            )
        )
        student_ids = [participant.student_id for participant in participants]

        checkins = {}
        for checkin in activity.checkins.filter(
            student_id__in=student_ids,
        ).order_by('student_id', '-check_in_time'):
            checkins.setdefault(checkin.student_id, checkin.check_in_time)

        checkouts = {}
        for checkout in activity.checkouts.filter(
            student_id__in=student_ids,
        ).order_by('student_id', '-check_out_time'):
            checkouts.setdefault(checkout.student_id, checkout.check_out_time)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Danh sách tham gia'

        headers = [
            'Mã sinh viên',
            'Họ và tên',
            'Thời gian check-in',
            'Thời gian check-out',
            'Trạng thái',
        ]
        sheet.append(headers)

        header_fill = PatternFill('solid', fgColor='1D4ED8')
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for participant in participants:
            student = participant.student
            check_in_time = checkins.get(student.pk)
            check_out_time = checkouts.get(student.pk)
            sheet.append([
                student.student_id,
                student.full_name,
                timezone.localtime(check_in_time).replace(tzinfo=None) if check_in_time else None,
                timezone.localtime(check_out_time).replace(tzinfo=None) if check_out_time else None,
                participant.get_status_display(),
            ])

        for row in sheet.iter_rows(min_row=2, min_col=3, max_col=4):
            for cell in row:
                cell.number_format = 'dd/mm/yyyy hh:mm:ss'

        sheet.freeze_panes = 'A2'
        sheet.auto_filter.ref = sheet.dimensions
        sheet.column_dimensions['A'].width = 18
        sheet.column_dimensions['B'].width = 32
        sheet.column_dimensions['C'].width = 24
        sheet.column_dimensions['D'].width = 24
        sheet.column_dimensions['E'].width = 20

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = (
            f'attachment; filename="activity_{activity.pk}_participants.xlsx"'
        )
        workbook.save(response)
        return response

    @action(
        detail=True,
        methods=['delete'],
        url_path='cancel-registration',
        permission_classes=[permissions.IsAuthenticated],
    )
    def cancel_registration(self, request, pk=None):
        activity = self.get_object()
        student = getattr(request.user, 'student_profile', None)
        if not student:
            return Response(
                {'error': 'Tài khoản không có hồ sơ sinh viên.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        participant = get_object_or_404(
            ActivityParticipant,
            activity=activity,
            student=student,
        )
        if participant.status != 'registered':
            return Response(
                {'error': 'Chỉ có thể hủy khi đăng ký đang ở trạng thái đã đăng ký.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if ActivityCheckIn.objects.filter(
            activity=activity,
            student=student,
        ).exists():
            return Response(
                {'error': 'Không thể hủy đăng ký sau khi đã check-in.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        cancellation_deadline = _activity_start_at(activity) - datetime.timedelta(
            hours=24,
        )
        if timezone.now() > cancellation_deadline:
            return Response(
                {'error': 'Chỉ có thể hủy đăng ký trước khi hoạt động diễn ra ít nhất 24 giờ.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        participant.delete()
        return Response({
            'message': 'Hủy đăng ký hoạt động thành công.',
            'activity': ActivitySerializer(activity, context={'request': request}).data,
        })

    @action(detail=True, methods=['post'], url_path='register')
    def register(self, request, pk=None):
        activity = self.get_object()
        student_id = request.data.get('studentId') or request.data.get('student_id')
        student = get_object_or_404(Student, student_id=student_id)
        
        # 1. Validate Scope
        if activity.scope_type == 'class':
            if not student.class_info or not activity.allowed_classes.filter(id=student.class_info.id).exists():
                return Response({'error': 'Lớp của bạn không nằm trong phạm vi tham gia hoạt động này.'}, status=status.HTTP_400_BAD_REQUEST)
        elif activity.scope_type == 'club':
            if not student.user or not student.user.user_organizations.filter(organization__in=activity.allowed_clubs.all()).exists():
                return Response({'error': 'Bạn không thuộc câu lạc bộ (CLB) được phép tham gia hoạt động này.'}, status=status.HTTP_400_BAD_REQUEST)

        # 2. Validate Pre-registration Time
        if activity.is_registration_required:
            now = timezone.now()
            if activity.registration_start and now < activity.registration_start:
                return Response({'error': 'Thời gian đăng ký tham gia hoạt động chưa bắt đầu.'}, status=status.HTTP_400_BAD_REQUEST)
            if activity.registration_end and now > activity.registration_end:
                return Response({'error': 'Thời gian đăng ký tham gia hoạt động đã kết thúc.'}, status=status.HTTP_400_BAD_REQUEST)

        participant, created, capacity_full = _reserve_activity_participant(
            activity.pk,
            student,
        )
        if capacity_full:
            return Response(
                {'error': 'Hoạt động đã đủ số lượng người tham gia tối đa.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if created:
            create_notification(
                user=request.user,
                title="Đăng ký hoạt động thành công",
                message=f"Bạn đã đăng ký tham gia hoạt động '{activity.title}' thành công.",
                type='activity',
                level='success',
                action_url='/'
            )
            if student.email:
                try:
                    from django.core.mail import send_mail
                    from django.conf import settings
                    
                    time_str = f"{activity.start_time.strftime('%H:%M') if activity.start_time else '00:00'} ngày {activity.date.strftime('%d/%m/%Y') if activity.date else 'chưa xác định'}"
                    location_str = activity.location or "Hội trường/Online"
                    points_str = str(activity.points)
                    
                    email_subject = f"[ITC Point] Đăng ký hoạt động thành công: {activity.title}"
                    email_message = f"Chào {student.full_name},\n\nHệ thống xác nhận bạn đã đăng ký tham gia hoạt động '{activity.title}' thành công.\n\nThời gian: {time_str}\nĐịa điểm: {location_str}\nĐiểm rèn luyện: +{points_str}đ."
                    
                    email_html = f"""
<div style="font-family: 'Segoe UI', Roboto, Helvetica, Arial, sans-serif; max-width: 600px; margin: 0 auto; border: 1px solid #e2e8f0; border-radius: 16px; overflow: hidden; box-shadow: 0 10px 25px -5px rgba(0,0,0,0.05), 0 8px 10px -6px rgba(0,0,0,0.05); background-color: #ffffff;">
  <div style="background: linear-gradient(135deg, #10b981, #059669); padding: 32px 24px; text-align: center; color: white;">
    <h1 style="margin: 0; font-size: 24px; font-weight: 800; letter-spacing: -0.5px; text-shadow: 0 1px 2px rgba(0,0,0,0.1);">Đăng Ký Hoạt Động</h1>
    <p style="margin: 4px 0 0 0; font-size: 13px; color: rgba(255,255,255,0.85); font-weight: 500;">Ghi nhận đăng ký tham gia thành công</p>
  </div>
  <div style="padding: 40px 32px; background-color: #ffffff;">
    <h2 style="margin-top: 0; color: #1e293b; font-size: 20px; font-weight: 700; letter-spacing: -0.5px;">Chào {student.full_name},</h2>
    <p style="color: #475569; line-height: 1.6; font-size: 15px; margin-top: 12px;">Hệ thống xác nhận bạn đã đăng ký tham gia hoạt động dưới đây thành công:</p>
    
    <div style="background-color: #f8fafc; border-radius: 12px; padding: 20px; margin: 28px 0; border: 1px solid #f1f5f9; border-left: 4px solid #10b981;">
      <table style="width: 100%; border-collapse: collapse;">
        <tr>
          <td style="padding: 8px 0; font-size: 14px; color: #64748b; font-weight: 600; width: 130px; text-transform: uppercase; letter-spacing: 0.5px;">Hoạt động:</td>
          <td style="padding: 8px 0; font-size: 15px; color: #0f172a; font-weight: 700;">{activity.title}</td>
        </tr>
        <tr>
          <td style="padding: 8px 0; font-size: 14px; color: #64748b; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px;">Thời gian:</td>
          <td style="padding: 8px 0; font-size: 15px; color: #0f172a; font-weight: 500;">{time_str}</td>
        </tr>
        <tr>
          <td style="padding: 8px 0; font-size: 14px; color: #64748b; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px;">Địa điểm:</td>
          <td style="padding: 8px 0; font-size: 15px; color: #0f172a; font-weight: 500;">{location_str}</td>
        </tr>
        <tr>
          <td style="padding: 8px 0; font-size: 14px; color: #64748b; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px;">Điểm cộng:</td>
          <td style="padding: 8px 0; font-size: 16px; color: #10b981; font-weight: 700;">+{points_str}đ rèn luyện</td>
        </tr>
      </table>
    </div>

    <div style="background-color: #eff6ff; border-radius: 8px; padding: 14px 18px; border: 1px solid #dbeafe; margin-top: 24px;">
      <p style="color: #1e40af; font-size: 13px; font-weight: 600; margin: 0; line-height: 1.5;">
        * Nhắc nhở: Hãy có mặt đúng giờ và thực hiện điểm danh Check-in khi bắt đầu, Check-out khi kết thúc hoạt động nhằm tích lũy điểm rèn luyện.
      </p>
    </div>
  </div>
  <div style="background-color: #f8fafc; padding: 24px; text-align: center; border-top: 1px solid #f1f5f9; color: #94a3b8; font-size: 12px; line-height: 1.5;">
    <p style="margin: 0 0 6px 0; font-weight: 500;">Email này được hệ thống ITC Point gửi tự động.</p>
    <p style="margin: 0;">© 2026 ITC Point. All rights reserved.</p>
  </div>
</div>
"""
                    send_mail(
                        email_subject,
                        email_message,
                        getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@itcpoint.com'),
                        [student.email],
                        fail_silently=False,
                        html_message=email_html
                    )
                except Exception as email_err:
                    print(f"Error sending registration confirmation email to {student.student_id}: {email_err}")
        return Response(ActivitySerializer(activity).data)

    @action(detail=True, methods=['post'], url_path='submit-evidence')
    def submit_evidence(self, request, pk=None):
        activity = self.get_object()
        student_id = request.data.get('studentId') or request.data.get('student_id')
        evidence_url = request.data.get('evidenceUrl')
        student = get_object_or_404(Student, student_id=student_id)
        
        participant = get_object_or_404(ActivityParticipant, activity=activity, student=student)
        participant.status = 'evidence_submitted'
        participant.evidence_url = evidence_url
        participant.save()

        # Fraud Rule 8 check: If modifying points/status after CTSV approval, log it
        # (For this mock, we write an audit log entry)
        AuditLog.objects.create(
            user=request.user if request.user.is_authenticated else None,
            action="Nộp minh chứng hoạt động",
            entity_name="ActivityParticipant",
            entity_id=participant.id,
            before_value="registered",
            after_value="evidence_submitted",
            ip_address=request.META.get('REMOTE_ADDR')
        )
        return Response(ActivitySerializer(activity).data)

    @action(detail=True, methods=['post'], url_path='confirm-attended')
    def confirm_attended(self, request, pk=None):
        activity = self.get_object()
        student_id = request.data.get('studentId') or request.data.get('student_id')
        student = get_object_or_404(Student, student_id=student_id)
        participant = get_object_or_404(ActivityParticipant, activity=activity, student=student)
        participant.status = 'attended'
        participant.save()
        return Response(ActivitySerializer(activity).data)

    @action(detail=True, methods=['post'], url_path='reject-evidence')
    def reject_evidence(self, request, pk=None):
        activity = self.get_object()
        student_id = request.data.get('studentId') or request.data.get('student_id')
        comment = request.data.get('comment', 'Minh chứng không hợp lệ.')
        student = get_object_or_404(Student, student_id=student_id)
        participant = get_object_or_404(ActivityParticipant, activity=activity, student=student)
        participant.status = 'registered'
        participant.evidence_url = None
        participant.save()
        
        # Send email notification
        if student.email:
            try:
                from django.core.mail import send_mail
                from django.conf import settings
                email_subject = f"[ITC Point] Từ chối minh chứng hoạt động: {activity.title}"
                email_message = f"Chào {student.full_name},\n\nMinh chứng tham gia hoạt động '{activity.title}' của bạn đã bị từ chối.\n\nLý do: {comment}\n\nVui lòng đăng nhập hệ thống để nộp lại minh chứng hợp lệ."
                send_mail(
                    email_subject,
                    email_message,
                    getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@itcpoint.com'),
                    [student.email],
                    fail_silently=True
                )
            except Exception as e:
                print(f"Error sending reject evidence email: {e}")
                
        return Response(ActivitySerializer(activity).data)

    @action(detail=True, methods=['post'], url_path='import-attendance')
    def import_attendance(self, request, pk=None):
        activity = self.get_object()
        student_codes = request.data.get('student_codes', [])
        if not isinstance(student_codes, list):
            return Response({'error': 'Danh sách mã sinh viên không hợp lệ.'}, status=status.HTTP_400_BAD_REQUEST)
        
        success_count = 0
        failed_codes = []
        
        for code in student_codes:
            code_str = str(code).strip()
            if not code_str:
                continue
            try:
                student = Student.objects.filter(student_id__iexact=code_str).first()
                if not student:
                    failed_codes.append(code_str)
                    continue
                
                from .models import ActivityParticipant
                participant, created = ActivityParticipant.objects.get_or_create(
                    activity=activity,
                    student=student,
                    defaults={'status': 'attended'}
                )
                if not created:
                    participant.status = 'attended'
                    participant.save()
                success_count += 1
            except Exception:
                failed_codes.append(code_str)
                
        return Response({
            'message': f'Nhập danh sách điểm danh thành công: {success_count} sinh viên.',
            'success_count': success_count,
            'failed_codes': failed_codes
        })


    @action(detail=True, methods=['post'], url_path='approve-points')
    def approve_points(self, request, pk=None):
        activity = self.get_object()
        activity.status = 'completed'
        activity.save()
        return Response(ActivitySerializer(activity).data)

    @action(detail=True, methods=['post'], url_path='check-in', permission_classes=[permissions.IsAuthenticated])
    def check_in(self, request, pk=None):
        activity = self.get_object()
        if not hasattr(request.user, 'student_profile') or not request.user.student_profile:
            return Response({'error': 'Tài khoản không phải là sinh viên hoặc không có hồ sơ sinh viên'}, status=status.HTTP_400_BAD_REQUEST)
        student = request.user.student_profile

        # Check registration requirement
        has_registered = ActivityParticipant.objects.filter(
            activity=activity,
            student=student,
        ).exists()
        if activity.is_registration_required and not has_registered:
            return Response({'error': 'Bạn cần đăng ký trước để tham gia hoạt động này.'}, status=status.HTTP_400_BAD_REQUEST)

        # Auto-register student if registration is not required
        if not activity.is_registration_required and not has_registered:
            _, _, capacity_full = _reserve_activity_participant(
                activity.pk,
                student,
            )
            if capacity_full:
                return Response(
                    {'error': 'Hoạt động đã đủ số lượng người tham gia tối đa.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        
        # Check check-in timing using the configured local event schedule.
        now = timezone.now()
        event_start, event_end = _activity_schedule(activity)
        checkin_start = event_start - datetime.timedelta(minutes=10)
        if now < checkin_start:
            return Response(
                {'error': 'Hoạt động chưa mở check-in. Vui lòng quay lại trước giờ diễn ra 10 phút.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if now > event_end:
            return Response(
                {'error': 'Hoạt động đã kết thúc, không thể check-in.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        face_verification, face_error = _verify_attendance_face(request)
        if face_error:
            return face_error
        location, location_error = _verify_attendance_location(request, activity, student)
        if location_error:
            return location_error

        device_id = request.data.get('deviceId') or request.data.get('device_id', 'unknown_device')
        ip_addr = request.data.get('ipAddress') or request.data.get('ip_address') or request.META.get('REMOTE_ADDR', '127.0.0.1')

        existing_checkin = ActivityCheckIn.objects.filter(activity=activity, student=student).first()
        if existing_checkin:
            return Response({
                'message': 'Bạn đã check-in hoạt động này.',
                'face_verified': True,
                'face_similarity': face_verification['similarity'],
                'gps_valid': True,
                'distance_meters': location['distance'],
                'check_in': ActivityCheckInSerializer(existing_checkin).data
            })

        # Keep the shared-device anomaly rule as an additional signal.
        three_minutes_ago = timezone.now() - timezone.timedelta(minutes=3)
        shared_device_checkins = ActivityCheckIn.objects.filter(
            activity=activity,
            device_id=device_id,
            check_in_time__gte=three_minutes_ago
        ).exclude(student=student)

        if shared_device_checkins.exists():
            other_students = ", ".join([ci.student.student_id for ci in shared_device_checkins])
            FraudDetection.objects.create(
                student=student,
                activity=activity,
                rule_code="RULE_5",
                severity="High",
                description=f"Nhiều tài khoản check-in chung một thiết bị ({device_id}) trong thời gian ngắn: {other_students}"
            )
            create_notification(
                user=request.user,
                title="Cảnh báo nghi ngờ gian lận điểm danh",
                message=f"Hệ thống phát hiện thiết bị của bạn ({device_id}) được dùng để điểm danh cho nhiều tài khoản khác nhau trong thời gian ngắn.",
                type='warning',
                level='danger',
                action_url='/'
            )

        # Save check-in
        checkin_obj = ActivityCheckIn.objects.create(
            activity=activity,
            student=student,
            latitude=location['latitude'],
            longitude=location['longitude'],
            gps_accuracy=location['accuracy'],
            face_similarity=face_verification['similarity'],
            face_liveness=face_verification['liveness'],
            face_realness=face_verification['realness'],
            device_id=device_id,
            ip_address=ip_addr
        )

        # Send notification to student
        create_notification(
            user=request.user,
            title="Điểm danh vào sự kiện thành công",
            message=f"Bạn đã điểm danh thành công vào hoạt động '{activity.title}' bằng Face ID.",
            type='activity',
            level='success',
            action_url='/'
        )

        # Initialize/update Attendance record
        attendance, _ = ActivityAttendance.objects.get_or_create(activity=activity, student=student)
        attendance.save()

        # Send confirmation email
        if student.email:
            try:
                from django.core.mail import send_mail
                from django.conf import settings
                local_time = timezone.localtime(checkin_obj.check_in_time)
                time_str = local_time.strftime("%H:%M:%S %d/%m/%Y")
                
                email_subject = f"[ITC Point] Check-in thành công: {activity.title}"
                email_message = f"Chào {student.full_name},\n\nHệ thống ghi nhận bạn đã Check-in bằng quét khuôn mặt (Face ID) thành công vào hoạt động '{activity.title}' vào lúc {time_str}.\n\nĐừng quên thực hiện Check-out khi hoạt động kết thúc để được tích lũy điểm rèn luyện."
                email_html = f"""
<div style="font-family: 'Segoe UI', Roboto, Helvetica, Arial, sans-serif; max-width: 600px; margin: 0 auto; border: 1px solid #e2e8f0; border-radius: 16px; overflow: hidden; box-shadow: 0 10px 25px -5px rgba(0,0,0,0.05); background-color: #ffffff;">
  <div style="background: linear-gradient(135deg, #10b981, #059669); padding: 32px 24px; text-align: center; color: white;">
    <h1 style="margin: 0; font-size: 24px; font-weight: 800;">Check-in Thành Công</h1>
    <p style="margin: 4px 0 0 0; font-size: 13px; color: rgba(255,255,255,0.85); font-weight: 500;">Xác thực bằng Face ID thành công</p>
  </div>
  <div style="padding: 40px 32px; background-color: #ffffff;">
    <h2 style="margin-top: 0; color: #1e293b; font-size: 20px; font-weight: 700;">Chào {student.full_name},</h2>
    <p style="color: #475569; line-height: 1.6; font-size: 15px;">Hệ thống xác nhận bạn đã thực hiện Check-in thành công cho hoạt động ngoại khóa:</p>
    
    <div style="background-color: #f0fdf4; border-radius: 12px; padding: 20px; margin: 28px 0; border: 1px solid #dcfce7; border-left: 4px solid #10b981;">
      <table style="width: 100%; border-collapse: collapse;">
        <tr>
          <td style="padding: 8px 0; font-size: 14px; color: #64748b; font-weight: 600; width: 150px; text-transform: uppercase;">Hoạt động:</td>
          <td style="padding: 8px 0; font-size: 15px; color: #0f172a; font-weight: 700;">{activity.title}</td>
        </tr>
        <tr>
          <td style="padding: 8px 0; font-size: 14px; color: #64748b; font-weight: 600; text-transform: uppercase;">Thời gian Check-in:</td>
          <td style="padding: 8px 0; font-size: 15px; color: #0f172a; font-weight: 500;">{time_str}</td>
        </tr>
        <tr>
          <td style="padding: 8px 0; font-size: 14px; color: #64748b; font-weight: 600; text-transform: uppercase;">Xác thực Face ID:</td>
          <td style="padding: 8px 0; font-size: 15px; color: #10b981; font-weight: 700;">Khớp khuôn mặt ({int(face_verification['similarity'] * 100)}%)</td>
        </tr>
      </table>
    </div>

    <div style="background-color: #fef3c7; border-radius: 8px; padding: 14px 18px; border: 1px solid #fde68a;">
      <p style="color: #b45309; font-size: 13px; font-weight: 600; margin: 0; line-height: 1.5;">
        * Nhắc nhở quan trọng: Đừng quên thực hiện Check-out khi hoạt động kết thúc để được hệ thống tự động ghi nhận điểm rèn luyện (+{activity.points}đ).
      </p>
    </div>
  </div>
  <div style="background-color: #f8fafc; padding: 24px; text-align: center; border-top: 1px solid #f1f5f9; color: #94a3b8; font-size: 12px;">
    <p style="margin: 0 0 6px 0; font-weight: 500;">Email này được hệ thống ITC Point gửi tự động.</p>
    <p style="margin: 0;">© 2026 ITC Point. All rights reserved.</p>
  </div>
</div>
"""
                send_mail(
                    email_subject,
                    email_message,
                    getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@itcpoint.com'),
                    [student.email],
                    fail_silently=True,
                    html_message=email_html
                )
            except Exception as e:
                print(f"Error sending check-in confirmation email: {e}")

        return Response({
            'message': 'Check-in thành công',
            'face_verified': True,
            'face_similarity': face_verification['similarity'],
            'gps_valid': True,
            'distance_meters': location['distance'],
            'gps_accuracy': location['accuracy'],
            'check_in': ActivityCheckInSerializer(checkin_obj).data
        })

    @action(detail=True, methods=['post'], url_path='check-out', permission_classes=[permissions.IsAuthenticated])
    def check_out(self, request, pk=None):
        activity = self.get_object()
        if not hasattr(request.user, 'student_profile') or not request.user.student_profile:
            return Response({'error': 'Tài khoản không phải là sinh viên hoặc không có hồ sơ sinh viên'}, status=status.HTTP_400_BAD_REQUEST)
        student = request.user.student_profile

        # Validate that check-in occurred
        checkin_obj = ActivityCheckIn.objects.filter(activity=activity, student=student).order_by('-check_in_time').first()
        if not checkin_obj:
            return Response({'error': 'Bạn chưa thực hiện check-in cho hoạt động này.'}, status=status.HTTP_400_BAD_REQUEST)

        # Check-out opens after two thirds of the scheduled event duration.
        now = timezone.now()
        event_start, event_end = _activity_schedule(activity)
        checkout_start = event_start + (event_end - event_start) * (2 / 3)
        checkout_end = event_end + timezone.timedelta(hours=1)
        if now < checkout_start:
            remaining_seconds = int((checkout_start - now).total_seconds())
            remaining_mins = max(1, math.ceil(remaining_seconds / 60))
            return Response({
                'error': (
                    'Chưa thể check-out. Hoạt động phải trôi qua ít nhất 2/3 '
                    f'thời lượng (còn khoảng {remaining_mins} phút nữa).'
                ),
            }, status=status.HTTP_400_BAD_REQUEST)
        if now > checkout_end:
            return Response({
                'error': 'Đã quá thời gian check-out cho phép (tối đa 1 tiếng sau khi hoạt động kết thúc).'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        face_verification, face_error = _verify_attendance_face(request)
        if face_error:
            return face_error
        location, location_error = _verify_attendance_location(request, activity, student)
        if location_error:
            return location_error

        device_id = request.data.get('deviceId') or request.data.get('device_id', 'unknown_device')
        ip_addr = request.data.get('ipAddress') or request.data.get('ip_address') or request.META.get('REMOTE_ADDR', '127.0.0.1')

        # Save check-out
        checkout_obj = ActivityCheckOut.objects.create(
            activity=activity,
            student=student,
            latitude=location['latitude'],
            longitude=location['longitude'],
            gps_accuracy=location['accuracy'],
            face_similarity=face_verification['similarity'],
            face_liveness=face_verification['liveness'],
            face_realness=face_verification['realness'],
            device_id=device_id,
            ip_address=ip_addr
        )

        # Calculate duration
        checkin_obj = ActivityCheckIn.objects.filter(activity=activity, student=student).order_by('-check_in_time').first()
        duration_mins = 0
        completion_pct = 0.0
        is_completed = False

        if checkin_obj:
            delta = checkout_obj.check_out_time - checkin_obj.check_in_time
            duration_mins = max(0, int(delta.total_seconds() / 60))

        event_duration_seconds = (event_end - event_start).total_seconds()
        elapsed_event_seconds = (checkout_obj.check_out_time - event_start).total_seconds()
        if event_duration_seconds > 0:
            completion_pct = max(
                0.0,
                min(100.0, elapsed_event_seconds / event_duration_seconds * 100),
            )
        is_completed = checkout_obj.check_out_time >= checkout_start

        # Update Attendance
        attendance, _ = ActivityAttendance.objects.get_or_create(activity=activity, student=student)
        attendance.duration_minutes = duration_mins
        attendance.completion_percent = completion_pct
        attendance.is_completed = is_completed
        attendance.save()

        # Update Participant status to 'attended' if completed
        if is_completed:
            participant, _ = ActivityParticipant.objects.get_or_create(
                activity=activity, student=student,
                defaults={'status': 'attended'}
            )
            if participant.status != 'attended':
                participant.status = 'attended'
                participant.save()
            create_notification(
                user=request.user,
                title="Điểm danh ra sự kiện thành công",
                message=f"Bạn đã điểm danh ra và hoàn thành hoạt động '{activity.title}' với thời gian tham gia {duration_mins} phút.",
                type='activity',
                level='success',
                action_url='/'
            )
            if student.email:
                try:
                    from django.core.mail import send_mail
                    from django.conf import settings
                    local_time = timezone.localtime(checkout_obj.check_out_time)
                    time_str = local_time.strftime("%H:%M:%S %d/%m/%Y")
                    
                    email_subject = f"[ITC Point] Check-out thành công: {activity.title}"
                    email_message = f"Chào {student.full_name},\n\nHệ thống ghi nhận bạn đã Check-out bằng quét khuôn mặt (Face ID) thành công hoạt động '{activity.title}' vào lúc {time_str}.\n\nThời gian tham gia: {duration_mins} phút ({int(completion_pct)}% thời lượng hoạt động).\n\nBạn được tích lũy +{activity.points}đ điểm rèn luyện từ hoạt động này."
                    email_html = f"""
<div style="font-family: 'Segoe UI', Roboto, Helvetica, Arial, sans-serif; max-width: 600px; margin: 0 auto; border: 1px solid #e2e8f0; border-radius: 16px; overflow: hidden; box-shadow: 0 10px 25px -5px rgba(0,0,0,0.05); background-color: #ffffff;">
  <div style="background: linear-gradient(135deg, #10b981, #059669); padding: 32px 24px; text-align: center; color: white;">
    <h1 style="margin: 0; font-size: 24px; font-weight: 800;">Check-out Thành Công</h1>
    <p style="margin: 4px 0 0 0; font-size: 13px; color: rgba(255,255,255,0.85); font-weight: 500;">Hoàn thành hoạt động và nhận điểm rèn luyện</p>
  </div>
  <div style="padding: 40px 32px; background-color: #ffffff;">
    <h2 style="margin-top: 0; color: #1e293b; font-size: 20px; font-weight: 700;">Chào {student.full_name},</h2>
    <p style="color: #475569; line-height: 1.6; font-size: 15px;">Chúc mừng bạn đã hoàn thành hoạt động ngoại khóa sau:</p>
    
    <div style="background-color: #f0fdf4; border-radius: 12px; padding: 20px; margin: 28px 0; border: 1px solid #dcfce7; border-left: 4px solid #10b981;">
      <table style="width: 100%; border-collapse: collapse;">
        <tr>
          <td style="padding: 8px 0; font-size: 14px; color: #64748b; font-weight: 600; width: 150px; text-transform: uppercase;">Hoạt động:</td>
          <td style="padding: 8px 0; font-size: 15px; color: #0f172a; font-weight: 700;">{activity.title}</td>
        </tr>
        <tr>
          <td style="padding: 8px 0; font-size: 14px; color: #64748b; font-weight: 600; text-transform: uppercase;">Thời gian Check-out:</td>
          <td style="padding: 8px 0; font-size: 15px; color: #0f172a; font-weight: 500;">{time_str}</td>
        </tr>
        <tr>
          <td style="padding: 8px 0; font-size: 14px; color: #64748b; font-weight: 600; text-transform: uppercase;">Thời gian tham gia:</td>
          <td style="padding: 8px 0; font-size: 15px; color: #0f172a; font-weight: 500;">{duration_mins} phút ({int(completion_pct)}% thời lượng)</td>
        </tr>
        <tr>
          <td style="padding: 8px 0; font-size: 14px; color: #64748b; font-weight: 600; text-transform: uppercase;">Điểm DRL cộng:</td>
          <td style="padding: 8px 0; font-size: 16px; color: #10b981; font-weight: 700;">+{activity.points} điểm</td>
        </tr>
      </table>
    </div>
  </div>
  <div style="background-color: #f8fafc; padding: 24px; text-align: center; border-top: 1px solid #f1f5f9; color: #94a3b8; font-size: 12px;">
    <p style="margin: 0 0 6px 0; font-weight: 500;">Email này được hệ thống ITC Point gửi tự động.</p>
    <p style="margin: 0;">© 2026 ITC Point. All rights reserved.</p>
  </div>
</div>
"""
                    send_mail(
                        email_subject,
                        email_message,
                        getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@itcpoint.com'),
                        [student.email],
                        fail_silently=True,
                        html_message=email_html
                    )
                except Exception as e:
                    print(f"Error sending check-out confirmation email: {e}")

        return Response({
            'message': 'Check-out thành công',
            'face_verified': True,
            'face_similarity': face_verification['similarity'],
            'gps_valid': True,
            'distance_meters': location['distance'],
            'gps_accuracy': location['accuracy'],
            'duration_minutes': duration_mins,
            'completion_percent': completion_pct,
            'is_completed': is_completed,
            'check_out': ActivityCheckOutSerializer(checkout_obj).data
        })


import string
import random

def generate_random_password(length=8):
    characters = string.ascii_letters + string.digits
    return ''.join(random.choice(characters) for i in range(length))

class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [permissions.AllowAny] # Restrict to admin in production

    def update(self, request, *args, **kwargs):
        target = self.get_object()
        changes_face = 'avatar' in request.data
        if changes_face:
            if not request.user.is_authenticated or (request.user.pk != target.pk and request.user.role != 'admin'):
                return Response(
                    {'error': 'Bạn không có quyền thay đổi dữ liệu Face ID của tài khoản này.'},
                    status=status.HTTP_403_FORBIDDEN,
                )
            
            avatar_base64 = request.data.get('avatar')
            if not avatar_base64:
                return Response({'error': 'Ảnh đại diện không được trống.'}, status=status.HTTP_400_BAD_REQUEST)
                
            embedding, error_msg = extract_face_embedding_from_base64(avatar_base64)
            if error_msg:
                return Response({'error': error_msg}, status=status.HTTP_400_BAD_REQUEST)
                
            # Copy data or mutate it to include avatar_embedding
            if isinstance(request.data, dict):
                request.data['avatar_embedding'] = embedding
            else:
                request.data._mutable = True
                request.data['avatar_embedding'] = embedding

        return super().update(request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        full_name = request.data.get('fullName') or request.data.get('full_name', '')
        email = request.data.get('email', '')
        student_id = request.data.get('studentId') or request.data.get('student_id', '')
        role = request.data.get('role', 'student')
        
        if not email:
            return Response({'error': 'Email is required'}, status=status.HTTP_400_BAD_REQUEST)

        # Auto-generate student_id if not provided
        if not student_id:
            ROLE_PREFIX_MAP = {
                'admin': 'AD',
                'advisor': 'GV',
                'student': 'SV',
                'class_monitor': 'SV',
                'organizer': 'BTC',
                'student_affairs': 'CTSV',
                'academic_affairs': 'DT',
            }
            prefix = ROLE_PREFIX_MAP.get(role, 'US')
            existing_ids = User.objects.filter(student_id__startswith=prefix).values_list('student_id', flat=True)
            max_num = 0
            for sid in existing_ids:
                if sid and sid.startswith(prefix):
                    num_part = sid[len(prefix):]
                    if num_part.isdigit():
                        max_num = max(max_num, int(num_part))
            next_num = max_num + 1
            student_id = f"{prefix}{str(next_num).zfill(4)}"
            
        username = student_id.lower()
        if User.objects.filter(username=username).exists() or User.objects.filter(student_id=student_id).exists():
            return Response({'error': f'Account with ID {student_id} already exists'}, status=status.HTTP_400_BAD_REQUEST)
            
        random_password = generate_random_password()
        
        user = User.objects.create_user(
            username=username,
            email=email,
            password=random_password,
            role=role,
            full_name=full_name,
            student_id=student_id,
            is_first_login=True,
            plain_password=random_password
        )
        
        # If student role, create a student profile
        if role in ['student', 'class_monitor']:
            Student.objects.create(
                user=user,
                student_id=student_id,
                full_name=full_name,
                email=email,
                faculty='Công nghệ Thông tin',
                cohort='K20'
            )
            
        serializer = self.get_serializer(user)
        return Response({
            'user': serializer.data,
            'password': random_password
        }, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], url_path='reset-password')
    def reset_password(self, request, pk=None):
        user = self.get_object()
        random_password = generate_random_password()
        user.set_password(random_password)
        user.is_first_login = True
        user.plain_password = random_password
        user.save()
        return Response({
            'message': 'Password reset successfully',
            'password': random_password
        })

    @action(detail=True, methods=['post'], url_path='toggle-active')
    def toggle_active(self, request, pk=None):
        user = self.get_object()
        user.is_active = not user.is_active
        user.save()
        return Response({
            'message': f"Account has been {'opened' if user.is_active else 'closed'} successfully",
            'is_active': user.is_active
        })

class OrganizationViewSet(viewsets.ModelViewSet):
    queryset = Organization.objects.prefetch_related('members').order_by('name')
    serializer_class = OrganizationSerializer
    permission_classes = [permissions.AllowAny]

    @transaction.atomic
    def perform_update(self, serializer):
        previous_name = serializer.instance.name
        organization = serializer.save()
        if organization.name != previous_name:
            Activity.objects.filter(organizer=previous_name).update(
                organizer=organization.name,
            )
            ExternalActivity.objects.filter(organizer_name=previous_name).update(
                organizer_name=organization.name,
            )

    def destroy(self, request, *args, **kwargs):
        organization = self.get_object()
        is_in_use = (
            organization.members.exists()
            or organization.activities.exists()
            or Activity.objects.filter(organizer=organization.name).exists()
            or ExternalActivity.objects.filter(
                organizer_name=organization.name,
            ).exists()
        )
        if is_in_use:
            return Response(
                {'detail': 'Không thể xóa đơn vị đang có thành viên hoặc hoạt động sử dụng.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return super().destroy(request, *args, **kwargs)

class UserOrganizationViewSet(viewsets.ModelViewSet):
    queryset = UserOrganization.objects.all()
    serializer_class = UserOrganizationSerializer
    permission_classes = [permissions.AllowAny]

class ClassPositionViewSet(viewsets.ModelViewSet):
    queryset = ClassPosition.objects.all()
    serializer_class = ClassPositionSerializer
    permission_classes = [permissions.AllowAny]

class StudentClassPositionViewSet(viewsets.ModelViewSet):
    queryset = StudentClassPosition.objects.all()
    serializer_class = StudentClassPositionSerializer
    permission_classes = [permissions.AllowAny]

class FraudDetectionViewSet(viewsets.ModelViewSet):
    queryset = FraudDetection.objects.all()
    serializer_class = FraudDetectionSerializer
    permission_classes = [permissions.AllowAny]

    def get_queryset(self):
        queryset = FraudDetection.objects.all()
        # support filters by rule_code, severity, student_id, activity_id
        rule_code = self.request.query_params.get('ruleCode') or self.request.query_params.get('rule_code')
        severity = self.request.query_params.get('severity')
        student = self.request.query_params.get('studentId') or self.request.query_params.get('student_id')
        activity = self.request.query_params.get('activityId') or self.request.query_params.get('activity_id')
        
        if rule_code:
            queryset = queryset.filter(rule_code=rule_code)
        if severity:
            queryset = queryset.filter(severity=severity)
        if student:
            queryset = queryset.filter(student__student_id=student)
        if activity:
            queryset = queryset.filter(activity_id=activity)
            
        return queryset

class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = AuditLog.objects.all()
    serializer_class = AuditLogSerializer
    permission_classes = [permissions.AllowAny]

class ChangeRequestViewSet(viewsets.ModelViewSet):
    queryset = ChangeRequest.objects.all()
    serializer_class = ChangeRequestSerializer
    permission_classes = [permissions.AllowAny]


class ExternalActivityViewSet(viewsets.ModelViewSet):
    serializer_class = ExternalActivitySerializer
    permission_classes = [permissions.AllowAny]

    def get_queryset(self):
        user = self.request.user
        queryset = ExternalActivity.objects.all().prefetch_related('evidence_files', 'fraud_flags', 'reviews')
        
        # Filtering by Student
        if user.is_authenticated and user.role == 'student':
            if hasattr(user, 'student_profile'):
                queryset = queryset.filter(student=user.student_profile)
            else:
                queryset = queryset.none()
        
        # Filtering by Advisor
        elif user.is_authenticated and user.role == 'advisor':
            queryset = queryset.filter(student__class_info__advisor=user)

        # Query parameter filters
        student_id = self.request.query_params.get('studentId')
        if student_id:
            queryset = queryset.filter(student__student_id=student_id)

        status_param = self.request.query_params.get('status')
        if status_param:
            queryset = queryset.filter(status=status_param)

        return queryset

    def perform_create(self, serializer):
        user = self.request.user
        if 'student' in serializer.validated_data:
            serializer.save()
        elif user.is_authenticated and hasattr(user, 'student_profile') and user.student_profile:
            serializer.save(student=user.student_profile)
        else:
            first_student = Student.objects.first()
            serializer.save(student=first_student)

    @action(detail=True, methods=['post'], url_path='submit')
    def submit_activity(self, request, pk=None):
        activity = self.get_object()
        if activity.status not in ['draft', 'need_more_info']:
            return Response({'error': 'Hoạt động phải ở trạng thái Draft hoặc Cần bổ sung mới được nộp.'}, status=status.HTTP_400_BAD_REQUEST)

        # Clear existing fraud flags for this activity
        activity.fraud_flags.all().delete()

        # Run Anti-fraud Rules Check
        evidence_files = activity.evidence_files.all()
        student = activity.student

        # Rule 1, 2, 3: Evidence File Hash Checks
        for ev in evidence_files:
            # Rule 1: Duplicate hash (any match in system)
            dup_any = EvidenceFile.objects.filter(file_hash=ev.file_hash).exclude(activity=activity)
            if dup_any.exists():
                FraudFlag.objects.create(
                    activity=activity,
                    rule_code='RULE_1',
                    severity='High',
                    description=f"Minh chứng '{ev.file_name}' trùng hash SHA256 với minh chứng khác trong hệ thống."
                )

                # Rule 2: One evidence used by multiple students
                dup_other_student = dup_any.exclude(activity__student=student)
                if dup_other_student.exists():
                    other_sv_ids = list(dup_other_student.values_list('activity__student__student_id', flat=True))
                    other_sv_ids_str = ", ".join(other_sv_ids)
                    FraudFlag.objects.create(
                        activity=activity,
                        rule_code='RULE_2',
                        severity='Critical',
                        description=f"Minh chứng '{ev.file_name}' được sử dụng bởi sinh viên khác (Mã SV: {other_sv_ids_str})."
                    )

                # Rule 3: Reused evidence (same student, different activities)
                dup_same_student = dup_any.filter(activity__student=student)
                if dup_same_student.exists():
                    other_act_names = list(dup_same_student.values_list('activity__activity_name', flat=True))
                    other_act_names_str = ", ".join(other_act_names)
                    FraudFlag.objects.create(
                        activity=activity,
                        rule_code='RULE_3',
                        severity='High',
                        description=f"Minh chứng '{ev.file_name}' bị nộp lại nhiều lần cho các hoạt động khác của bạn (Hoạt động: {other_act_names_str})."
                    )

        # Rule 4: Activity outside semester range
        # Typical semester break is July (7) and August (8)
        if activity.start_date.month in [7, 8] or activity.end_date.month in [7, 8]:
            FraudFlag.objects.create(
                activity=activity,
                rule_code='RULE_4',
                severity='Medium',
                description=f"Thời gian hoạt động ({activity.start_date} - {activity.end_date}) diễn ra ngoài khoảng thời gian học kỳ chính."
            )

        # Rule 5: Proposed score exceeds rules (> 10 points)
        if activity.proposed_score > 10:
            FraudFlag.objects.create(
                activity=activity,
                rule_code='RULE_5',
                severity='High',
                description=f"Số điểm đề xuất ({activity.proposed_score} điểm) vượt quá mức quy định cho hoạt động ngoài trường (tối đa 10 điểm)."
            )

        # Rule 6: Gained points for same activity name before
        gained_before = ExternalActivity.objects.filter(
            student=student,
            activity_name__iexact=activity.activity_name,
            status='approved'
        ).exclude(id=activity.id)
        if gained_before.exists():
            FraudFlag.objects.create(
                activity=activity,
                rule_code='RULE_6',
                severity='Critical',
                description=f"Sinh viên đã được cộng điểm cho hoạt động có cùng tên '{activity.activity_name}' trước đó."
            )

        # Rule 7: Organizer name in suspicious watchlist
        suspicious_keywords = ["tự phát", "không rõ", "cá nhân", "nhóm sinh viên", "nhóm tự phát", "scam", "chưa xác minh", "unknown"]
        org_name_lower = activity.organizer_name.lower()
        if any(kw in org_name_lower for kw in suspicious_keywords):
            FraudFlag.objects.create(
                activity=activity,
                rule_code='RULE_7',
                severity='Medium',
                description=f"Đơn vị tổ chức '{activity.organizer_name}' nằm trong danh sách cần xác minh hoặc có tên không rõ ràng."
            )

        # Update status
        activity.status = 'submitted'
        activity.save()

        # Audit logging
        AuditLog.objects.create(
            user=request.user if request.user.is_authenticated else None,
            action="Nộp hồ sơ hoạt động ngoài trường",
            entity_name="ExternalActivity",
            entity_id=activity.id,
            before_value="draft",
            after_value="submitted",
            ip_address=request.META.get('REMOTE_ADDR')
        )

        return Response(ExternalActivitySerializer(activity).data)

    @action(detail=True, methods=['post'], url_path='review-advisor')
    def review_advisor(self, request, pk=None):
        activity = self.get_object()
        reviewer = request.user if request.user.is_authenticated else None
        
        status_input = request.data.get('status') # 'advisor_approved', 'need_more_info', 'rejected_by_advisor'
        comment = request.data.get('comment', '')

        if status_input not in ['advisor_approved', 'need_more_info', 'rejected_by_advisor']:
            return Response({'error': 'Trạng thái xét duyệt của Cố vấn không hợp lệ.'}, status=status.HTTP_400_BAD_REQUEST)

        old_status = activity.status
        activity.status = status_input
        activity.save()

        # Create review entry
        EvidenceReview.objects.create(
            activity=activity,
            reviewer=reviewer,
            review_level='advisor',
            status=status_input,
            comment=comment
        )

        # Send notifications
        student_user = User.objects.filter(student_id=activity.student.student_id).first()
        if student_user:
            if status_input == 'advisor_approved':
                create_notification(
                    user=student_user,
                    title="Hoạt động ngoại khóa được Cố vấn thông qua",
                    message=f"Hoạt động '{activity.activity_name}' của bạn đã được Cố vấn học tập thông qua và chuyển lên phòng CTSV duyệt.",
                    type='activity',
                    level='success',
                    action_url='/'
                )
            elif status_input == 'need_more_info':
                create_notification(
                    user=student_user,
                    title="Yêu cầu bổ sung minh chứng hoạt động",
                    message=f"Cố vấn học tập yêu cầu bạn bổ sung thêm minh chứng cho hoạt động '{activity.activity_name}'. Nhận xét: {comment}",
                    type='activity',
                    level='warning',
                    action_url='/'
                )
                if activity.student.email:
                    try:
                        from django.core.mail import send_mail
                        from django.conf import settings
                        email_subject = f"[ITC Point] Yêu cầu bổ sung minh chứng hoạt động: {activity.activity_name}"
                        email_message = f"Chào {activity.student.full_name},\n\nCố vấn học tập yêu cầu bạn bổ sung thông tin minh chứng cho hoạt động ngoại khóa '{activity.activity_name}'.\n\nNhận xét: {comment}\n\nVui lòng đăng nhập hệ thống để chỉnh sửa và nộp lại."
                        email_html = f"""
<div style="font-family: 'Segoe UI', Roboto, Helvetica, Arial, sans-serif; max-width: 600px; margin: 0 auto; border: 1px solid #e2e8f0; border-radius: 16px; overflow: hidden; box-shadow: 0 10px 25px -5px rgba(0,0,0,0.05); background-color: #ffffff;">
  <div style="background: linear-gradient(135deg, #f59e0b, #d97706); padding: 32px 24px; text-align: center; color: white;">
    <h1 style="margin: 0; font-size: 24px; font-weight: 800;">Yêu Cầu Bổ Sung Minh Chứng</h1>
    <p style="margin: 4px 0 0 0; font-size: 13px; color: rgba(255,255,255,0.85); font-weight: 500;">{activity.activity_name}</p>
  </div>
  <div style="padding: 40px 32px; background-color: #ffffff;">
    <h2 style="margin-top: 0; color: #1e293b; font-size: 20px; font-weight: 700;">Chào {activity.student.full_name},</h2>
    <p style="color: #475569; line-height: 1.6; font-size: 15px;">Cố vấn học tập đã xem xét minh chứng hoạt động ngoại khóa của bạn và yêu cầu bổ sung thêm thông tin:</p>
    
    <div style="background-color: #fef3c7; border-radius: 12px; padding: 20px; margin: 28px 0; border: 1px solid #fde68a; border-left: 4px solid #f59e0b;">
      <p style="margin: 0 0 8px 0; font-size: 14px; color: #92400e; font-weight: 700; text-transform: uppercase;">Ý kiến phản hồi từ Cố vấn:</p>
      <p style="margin: 0; font-size: 15px; color: #78350f; font-style: italic;">"{comment}"</p>
    </div>

    <div style="text-align: center; margin: 36px 0 28px 0;">
      <a href="http://localhost:8080/" style="background: linear-gradient(135deg, #f59e0b, #d97706); color: #ffffff; padding: 14px 32px; text-decoration: none; border-radius: 10px; font-weight: 700; font-size: 15px; display: inline-block; box-shadow: 0 10px 15px -3px rgba(245, 158, 11, 0.3);">Cập nhật minh chứng</a>
    </div>
  </div>
  <div style="background-color: #f8fafc; padding: 24px; text-align: center; border-top: 1px solid #f1f5f9; color: #94a3b8; font-size: 12px;">
    <p style="margin: 0 0 6px 0; font-weight: 500;">Email này được hệ thống ITC Point gửi tự động.</p>
    <p style="margin: 0;">© 2026 ITC Point. All rights reserved.</p>
  </div>
</div>
"""
                        send_mail(
                            email_subject,
                            email_message,
                            getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@itcpoint.com'),
                            [activity.student.email],
                            fail_silently=True,
                            html_message=email_html
                        )
                    except Exception as e:
                        print(f"Error sending need_more_info email: {e}")
            elif status_input == 'rejected_by_advisor':
                create_notification(
                    user=student_user,
                    title="Minh chứng hoạt động bị Cố vấn từ chối",
                    message=f"Minh chứng hoạt động '{activity.activity_name}' đã bị Cố vấn học tập từ chối. Lý do: {comment}",
                    type='activity',
                    level='danger',
                    action_url='/'
                )
                if activity.student.email:
                    try:
                        from django.core.mail import send_mail
                        from django.conf import settings
                        email_subject = f"[ITC Point] Từ chối minh chứng hoạt động: {activity.activity_name}"
                        email_message = f"Chào {activity.student.full_name},\n\nMinh chứng hoạt động ngoại khóa '{activity.activity_name}' của bạn đã bị Cố vấn học tập từ chối.\n\nLý do từ chối: {comment}\n\nVui lòng đăng nhập hệ thống để biết thêm chi tiết."
                        email_html = f"""
<div style="font-family: 'Segoe UI', Roboto, Helvetica, Arial, sans-serif; max-width: 600px; margin: 0 auto; border: 1px solid #e2e8f0; border-radius: 16px; overflow: hidden; box-shadow: 0 10px 25px -5px rgba(0,0,0,0.05); background-color: #ffffff;">
  <div style="background: linear-gradient(135deg, #ef4444, #b91c1c); padding: 32px 24px; text-align: center; color: white;">
    <h1 style="margin: 0; font-size: 24px; font-weight: 800;">Minh Chứng Bị Từ Chối</h1>
    <p style="margin: 4px 0 0 0; font-size: 13px; color: rgba(255,255,255,0.85); font-weight: 500;">{activity.activity_name}</p>
  </div>
  <div style="padding: 40px 32px; background-color: #ffffff;">
    <h2 style="margin-top: 0; color: #1e293b; font-size: 20px; font-weight: 700;">Chào {activity.student.full_name},</h2>
    <p style="color: #475569; line-height: 1.6; font-size: 15px;">Chúng tôi rất tiếc phải thông báo minh chứng hoạt động ngoại khóa của bạn đã bị Cố vấn học tập từ chối:</p>
    
    <div style="background-color: #fef2f2; border-radius: 12px; padding: 20px; margin: 28px 0; border: 1px solid #fee2e2; border-left: 4px solid #ef4444;">
      <p style="margin: 0 0 8px 0; font-size: 14px; color: #991b1b; font-weight: 700; text-transform: uppercase;">Lý do từ chối:</p>
      <p style="margin: 0; font-size: 15px; color: #7f1d1d; font-style: italic;">"{comment}"</p>
    </div>

    <div style="text-align: center; margin: 36px 0 28px 0;">
      <a href="http://localhost:8080/" style="background: linear-gradient(135deg, #ef4444, #b91c1c); color: #ffffff; padding: 14px 32px; text-decoration: none; border-radius: 10px; font-weight: 700; font-size: 15px; display: inline-block; box-shadow: 0 10px 15px -3px rgba(239, 68, 68, 0.3);">Xem trên hệ thống</a>
    </div>
  </div>
  <div style="background-color: #f8fafc; padding: 24px; text-align: center; border-top: 1px solid #f1f5f9; color: #94a3b8; font-size: 12px;">
    <p style="margin: 0 0 6px 0; font-weight: 500;">Email này được hệ thống ITC Point gửi tự động.</p>
    <p style="margin: 0;">© 2026 ITC Point. All rights reserved.</p>
  </div>
</div>
"""
                        send_mail(
                            email_subject,
                            email_message,
                            getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@itcpoint.com'),
                            [activity.student.email],
                            fail_silently=True,
                            html_message=email_html
                        )
                    except Exception as e:
                        print(f"Error sending advisor rejection email: {e}")

        # Audit logging
        AuditLog.objects.create(
            user=reviewer,
            action="Cố vấn học tập xét duyệt hoạt động",
            entity_name="ExternalActivity",
            entity_id=activity.id,
            before_value=old_status,
            after_value=status_input,
            ip_address=request.META.get('REMOTE_ADDR')
        )

        return Response(ExternalActivitySerializer(activity).data)

    @action(detail=True, methods=['post'], url_path='review-ctsv')
    def review_ctsv(self, request, pk=None):
        activity = self.get_object()
        reviewer = request.user if request.user.is_authenticated else None
        
        status_input = request.data.get('status') # 'approved', 'rejected'
        comment = request.data.get('comment', '')

        if status_input not in ['approved', 'rejected']:
            return Response({'error': 'Trạng thái xét duyệt của CTSV không hợp lệ.'}, status=status.HTTP_400_BAD_REQUEST)

        old_status = activity.status
        activity.status = status_input
        activity.save()

        # Create review entry
        EvidenceReview.objects.create(
            activity=activity,
            reviewer=reviewer,
            review_level='ctsv',
            status=status_input,
            comment=comment
        )

        # Send notifications
        student_user = User.objects.filter(student_id=activity.student.student_id).first()
        if student_user:
            if status_input == 'approved':
                create_notification(
                    user=student_user,
                    title="Minh chứng hoạt động ngoại khóa được CTSV duyệt",
                    message=f"Minh chứng hoạt động ngoại khóa '{activity.activity_name}' đã được phòng CTSV duyệt thành công. Bạn được cộng {activity.proposed_score} điểm rèn luyện.",
                    type='activity',
                    level='success',
                    action_url='/'
                )
            elif status_input == 'rejected':
                create_notification(
                    user=student_user,
                    title="Minh chứng hoạt động ngoại khóa bị CTSV từ chối",
                    message=f"Minh chứng hoạt động ngoại khóa '{activity.activity_name}' đã bị phòng CTSV từ chối. Lý do: {comment}",
                    type='activity',
                    level='danger',
                    action_url='/'
                )
                if activity.student.email:
                    try:
                        from django.core.mail import send_mail
                        from django.conf import settings
                        email_subject = f"[ITC Point] Từ chối minh chứng hoạt động: {activity.activity_name}"
                        email_message = f"Chào {activity.student.full_name},\n\nMinh chứng hoạt động ngoại khóa '{activity.activity_name}' của bạn đã bị phòng CTSV từ chối.\n\nLý do từ chối: {comment}\n\nVui lòng đăng nhập hệ thống để biết thêm chi tiết."
                        email_html = f"""
<div style="font-family: 'Segoe UI', Roboto, Helvetica, Arial, sans-serif; max-width: 600px; margin: 0 auto; border: 1px solid #e2e8f0; border-radius: 16px; overflow: hidden; box-shadow: 0 10px 25px -5px rgba(0,0,0,0.05); background-color: #ffffff;">
  <div style="background: linear-gradient(135deg, #ef4444, #b91c1c); padding: 32px 24px; text-align: center; color: white;">
    <h1 style="margin: 0; font-size: 24px; font-weight: 800;">Minh Chứng Bị Từ Chối</h1>
    <p style="margin: 4px 0 0 0; font-size: 13px; color: rgba(255,255,255,0.85); font-weight: 500;">{activity.activity_name} (Phòng CTSV)</p>
  </div>
  <div style="padding: 40px 32px; background-color: #ffffff;">
    <h2 style="margin-top: 0; color: #1e293b; font-size: 20px; font-weight: 700;">Chào {activity.student.full_name},</h2>
    <p style="color: #475569; line-height: 1.6; font-size: 15px;">Chúng tôi rất tiếc phải thông báo minh chứng hoạt động ngoại khóa của bạn đã bị phòng CTSV từ chối:</p>
    
    <div style="background-color: #fef2f2; border-radius: 12px; padding: 20px; margin: 28px 0; border: 1px solid #fee2e2; border-left: 4px solid #ef4444;">
      <p style="margin: 0 0 8px 0; font-size: 14px; color: #991b1b; font-weight: 700; text-transform: uppercase;">Lý do từ chối:</p>
      <p style="margin: 0; font-size: 15px; color: #7f1d1d; font-style: italic;">"{comment}"</p>
    </div>

    <div style="text-align: center; margin: 36px 0 28px 0;">
      <a href="http://localhost:8080/" style="background: linear-gradient(135deg, #ef4444, #b91c1c); color: #ffffff; padding: 14px 32px; text-decoration: none; border-radius: 10px; font-weight: 700; font-size: 15px; display: inline-block; box-shadow: 0 10px 15px -3px rgba(239, 68, 68, 0.3);">Xem trên hệ thống</a>
    </div>
  </div>
  <div style="background-color: #f8fafc; padding: 24px; text-align: center; border-top: 1px solid #f1f5f9; color: #94a3b8; font-size: 12px;">
    <p style="margin: 0 0 6px 0; font-weight: 500;">Email này được hệ thống ITC Point gửi tự động.</p>
    <p style="margin: 0;">© 2026 ITC Point. All rights reserved.</p>
  </div>
</div>
"""
                        send_mail(
                            email_subject,
                            email_message,
                            getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@itcpoint.com'),
                            [activity.student.email],
                            fail_silently=True,
                            html_message=email_html
                        )
                    except Exception as e:
                        print(f"Error sending CTSV rejection email: {e}")

        # Audit logging
        AuditLog.objects.create(
            user=reviewer,
            action="Phòng CTSV phê duyệt cuối hoạt động ngoài trường",
            entity_name="ExternalActivity",
            entity_id=activity.id,
            before_value=old_status,
            after_value=status_input,
            ip_address=request.META.get('REMOTE_ADDR')
        )

        return Response(ExternalActivitySerializer(activity).data)

    @action(detail=False, methods=['post'], url_path='random-audit')
    def random_audit(self, request):
        # CTSV requests random audit of approved activities (usually 5% to 10%)
        import random
        percent = int(request.data.get('percent', 10))
        if percent < 5 or percent > 15:
            percent = 10 # Default to 10%

        approved_activities = list(ExternalActivity.objects.filter(status='approved'))
        total_count = len(approved_activities)
        if total_count == 0:
            return Response({'message': 'Không có hồ sơ đã duyệt nào để hậu kiểm.', 'audited_ids': []})

        audit_count = max(1, int(total_count * (percent / 100.0)))
        audited_samples = random.sample(approved_activities, audit_count)

        audited_ids = []
        for act in audited_samples:
            audited_ids.append(act.id)
            # Log the post-audit action
            AuditLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action="Hậu kiểm ngẫu nhiên hồ sơ hoạt động ngoài trường",
                entity_name="ExternalActivity",
                entity_id=act.id,
                before_value="approved",
                after_value="approved_audited",
                ip_address=request.META.get('REMOTE_ADDR')
            )

        return Response({
            'message': f'Đã chọn ngẫu nhiên {audit_count} hồ sơ ({percent}%) trên tổng số {total_count} hồ sơ đã duyệt để tiến hành hậu kiểm.',
            'audited_activities': ExternalActivitySerializer(audited_samples, many=True).data
        })


class EvidenceFileViewSet(viewsets.ModelViewSet):
    queryset = EvidenceFile.objects.all()
    serializer_class = EvidenceFileSerializer
    permission_classes = [permissions.AllowAny]

    def perform_create(self, serializer):
        # Calculate SHA256 file hash if not provided
        import hashlib
        file_obj = self.request.FILES.get('file')
        file_hash = self.request.data.get('file_hash')

        if file_obj and not file_hash:
            sha256 = hashlib.sha256()
            for chunk in file_obj.chunks():
                sha256.update(chunk)
            file_hash = sha256.hexdigest()
            serializer.save(file_hash=file_hash, file_size=file_obj.size)
        else:
            serializer.save()


class EvidenceReviewViewSet(viewsets.ModelViewSet):
    queryset = EvidenceReview.objects.all()
    serializer_class = EvidenceReviewSerializer
    permission_classes = [permissions.AllowAny]


class FraudFlagViewSet(viewsets.ModelViewSet):
    queryset = FraudFlag.objects.all()
    serializer_class = FraudFlagSerializer
    permission_classes = [permissions.AllowAny]


class SystemConfigViewSet(viewsets.ModelViewSet):
    queryset = SystemConfig.objects.all()
    serializer_class = SystemConfigSerializer
    lookup_field = 'key'
    permission_classes = [permissions.AllowAny]

    @action(detail=False, methods=['post'], url_path='set-value')
    def set_value(self, request):
        key = request.data.get('key')
        value = request.data.get('value')
        description = request.data.get('description', '')
        if not key:
            return Response({'detail': 'key is required.'}, status=status.HTTP_400_BAD_REQUEST)
        config, created = SystemConfig.objects.update_or_create(
            key=key,
            defaults={'value': value, 'description': description}
        )
        return Response({'success': True, 'key': config.key, 'value': config.value})

    def retrieve(self, request, *args, **kwargs):
        key = self.kwargs.get('key')
        if key == 'training_score_scale':
            config, created = SystemConfig.objects.get_or_create(
                key=key,
                defaults={
                    "value": [
                        {"min_score": 90, "label": "Xuất sắc"},
                        {"min_score": 80, "label": "Giỏi"},
                        {"min_score": 65, "label": "Khá"},
                        {"min_score": 50, "label": "Trung bình"},
                        {"min_score": 35, "label": "Yếu"},
                        {"min_score": 0, "label": "Kém"}
                    ],
                    "description": "Thang điểm xếp loại điểm rèn luyện"
                }
            )
            serializer = self.get_serializer(config)
            return Response(serializer.data)
        return super().retrieve(request, *args, **kwargs)


class NotificationViewSet(viewsets.ModelViewSet):
    serializer_class = NotificationSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Notification.objects.filter(user=self.request.user)

    @action(detail=True, methods=['post'], url_path='mark-as-read')
    def mark_as_read(self, request, pk=None):
        notification = self.get_object()
        notification.unread = False
        notification.save()
        return Response({'status': 'marked as read'})

    @action(detail=False, methods=['post'], url_path='mark-all-read')
    def mark_all_read(self, request):
        Notification.objects.filter(user=request.user, unread=True).update(unread=False)
        return Response({'status': 'all marked as read'})







